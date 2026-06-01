import numpy as np
import matplotlib.pyplot as plt


def set_plot_style(font='Times New Roman', math_fontset='stix', base_fontsize=11):
    """Set a global font (and a few related cosmetic defaults) for every
    matplotlib figure produced afterwards.

    Call this once at the top of your notebook — every subsequent
    `plt.plot(...)`, `fig.suptitle(...)`, axis label, legend etc. will pick
    up the new font.  Safe to call repeatedly.

    Args:
        font          : font family for non-math text.  Defaults to
                        'Times New Roman'.  Pass any installed font name.
                        Falls back silently if the family can't be resolved
                        (matplotlib will warn and use its default).
        math_fontset  : math/Greek glyph set; 'stix' pairs well with serif
                        text fonts. Other options: 'cm' (Computer Modern),
                        'dejavusans', etc.
        base_fontsize : default font size for text elements.  Title, axis
                        labels, legend and tick sizes are derived from this.

    Returns:
        The dict of rcParams keys that were touched, so you can `restore`
        them later if you want a localised override.
    """
    import matplotlib as mpl

    touched = {
        'font.family':           font,
        'font.serif':            [font, 'DejaVu Serif'],     # fallback chain
        'mathtext.fontset':      math_fontset,
        'mathtext.rm':           font,
        'mathtext.it':           f'{font}:italic',
        'mathtext.bf':           f'{font}:bold',
        'axes.titlesize':        base_fontsize + 2,
        'axes.labelsize':        base_fontsize,
        'xtick.labelsize':       base_fontsize - 1,
        'ytick.labelsize':       base_fontsize - 1,
        'legend.fontsize':       base_fontsize - 1,
        'figure.titlesize':      base_fontsize + 3,
        'font.size':             base_fontsize,
    }
    mpl.rcParams.update(touched)
    return touched


# --- plotting for GE framework with grid-based policies ---

def plot_policy(self, age=20, i_fix=0, i_z=0, i_housing=0, m_max_consumption=50,
                m_max_plot=None, m_max_choice=None, m_max_buffer=None,
                show_both_states=True, age_start=16):
	"""Plot consumption, buffer-stock target and choice-prob policies at a chosen age.

	Reads the age-specific policy from ``self.hh_policy`` (populated by
	``solve_hh_backwards``); ``self.sol.c``/``self.sol.a`` are not used here
	because the backward DC-EGM sweep overwrites them and only the age-0
	policy survives — which is essentially flat in m and not useful.

	Parameters
	----------
	age                : lifecycle index ``it`` (0 = newborn, J-1 = terminal)
	i_fix              : fixed-type index (kept for API compatibility)
	i_z                : productivity index used for income / PIH benchmark
	i_housing          : current housing state shown in the choice-prob panel
	                     and the state used for the PIH reference line
	                     (default 0 = rural renter, i.e. the "top-left" panel of
	                     ``plot_choice_probs``); also the state shown in the
	                     consumption / buffer panels when show_both_states=False
	m_max_consumption  : x-axis limit for the consumption panel (default 50)
	m_max_plot         : kept for backward compatibility (unused in 1×3 layout)
	m_max_choice       : x-axis limit for the choice-prob panel
	                     (default: ``m_max_consumption``)
	m_max_buffer       : x-axis limit for the buffer-stock panel — i.e. the
	                     largest ``a_{t-1}`` shown (default: ``m_max_consumption``)
	show_both_states   : if True, overlay all four current housing states in the
	                     consumption / buffer panels
	age_start          : real age corresponding to model age 0 (default 16)

	Layout (1×3)
	-----------
	(0) Consumption policy                          — x-axis = m, cap m_max_consumption
	(1) Buffer-stock change-in-assets diagram       — x-axis = a_{t-1}
	(2) Choice probabilities given current i_housing
	"""
	par = self.par
	_ = i_fix  # currently unused; kept for API compatibility

	if not hasattr(self, 'hh_policy'):
		raise AttributeError(
			"self.hh_policy not available — run solve_hh_backwards(model) first."
		)

	J_life = int(getattr(par, 'J', getattr(par, 'T', 0)))
	age = int(age)
	if age < 0 or age >= J_life:
		raise ValueError(f"age must be in [0, {J_life - 1}], got {age}")

	pol = self.hh_policy
	for i_h in range(4):
		if age not in pol.get(i_h, {}) or i_z not in pol[i_h][age]:
			raise KeyError(
				f"hh_policy missing entry for i_h={i_h}, age={age}, i_z={i_z}"
			)

	m_grid = par.m_grid if hasattr(par, 'm_grid') else par.a_grid
	m_lo      = 0.0
	m_hi_full = float(np.nanmax(m_grid)) if np.any(np.isfinite(m_grid)) else 1.0
	m_hi_cons = min(float(m_max_consumption), m_hi_full)
	m_hi_full_plot = m_hi_full if m_max_plot is None else min(float(m_max_plot), m_hi_full)
	m_hi_choice = (m_hi_cons if m_max_choice is None
	               else min(float(m_max_choice), m_hi_full))
	a_hi_buffer = (m_hi_cons if m_max_buffer is None
	               else min(float(m_max_buffer), m_hi_full))

	# Interpolate the endogenous-grid (c, a) policy at this (age, i_z) onto m_grid
	c_by_h = {}
	a_by_h = {}
	for i_h in range(4):
		pol_h = pol[i_h][age][i_z]
		c_endog = np.interp(m_grid, pol_h['m'], pol_h['c'])
		c_by_h[i_h] = c_endog
		a_by_h[i_h] = m_grid - c_endog

	# Interpolate the choice probabilities for the requested current state
	pol_cur = pol[i_housing][age][i_z]
	pr_endog = pol_cur['pr_choices']  # (4, N_endog)
	pr_reg = np.empty((4, m_grid.size))
	for k in range(4):
		pr_reg[k, :] = np.interp(m_grid, pol_cur['m'], pr_endog[k, :])
	pr_reg = np.clip(pr_reg, 0.0, 1.0)
	col_sums = pr_reg.sum(axis=0, keepdims=True)
	col_sums = np.where(col_sums > 0.0, col_sums, 1.0)
	pr_reg = pr_reg / col_sums

	# States shown in panels
	state_list  = [0, 1, 2, 3] if show_both_states else [i_housing]
	state_label = {0: 'rural renter (curr)', 1: 'urban renter (curr)',
	               2: 'rural owner (curr)',  3: 'urban owner (curr)'}
	state_color = {0: 'tab:blue', 1: 'tab:orange',
	               2: 'tab:green', 3: 'tab:red'}
	state_ls    = {0: '-', 1: '-', 2: '--', 3: '--'}
	next_labels = {0: 'rural renter', 1: 'urban renter',
	               2: 'rural owner',  3: 'urban owner'}
	curr_labels = {0: 'rural renter', 1: 'urban renter',
	               2: 'rural owner',  3: 'urban owner'}

	fig, axes = plt.subplots(1, 3, figsize=(16, 4.5))

	# ── Panel (0): Consumption policy ───────────────────────────────────────
	ax = axes[0]
	mask_cons = (m_grid >= m_lo) & (m_grid <= m_hi_cons)
	for i_h in state_list:
		c_pol = c_by_h[i_h]
		c_plot = np.where(np.isfinite(c_pol) & (c_pol > 0), c_pol, np.nan)
		ax.plot(m_grid[mask_cons], c_plot[mask_cons],
		        color=state_color[i_h], ls=state_ls[i_h], lw=2,
		        label=state_label[i_h])
	ax.plot(m_grid[mask_cons], m_grid[mask_cons], 'k--', alpha=0.4, lw=1,
	        label='45° line')
	ax.set_xlim(m_lo, m_hi_cons)
	ax.set_xlabel('Cash-on-hand (m)')
	ax.set_ylabel('Consumption (c)')
	ax.set_title(f'Consumption Policy  (m ≤ {m_hi_cons:g})')
	ax.legend(fontsize=9)
	ax.grid(True, alpha=0.4)

	# ── Panel (1): Buffer-stock change-in-assets diagram ────────────────────
	# x-axis: a_{t-1} (savings carried in); y-axis: a_t - a_{t-1}.
	# Where each line crosses 0 is the target savings for that current state.
	# Holds i_h_next = i_h_current (no transition) so the budget identity is
	#     m_t = R · a_{t-1} + w·exp(χ[t]+z) − housing_cost(i_h_current)
	#     a_t = m_t − c(m_t | i_h_current, age, i_z)
	# A perfect-foresight (PIH) benchmark is drawn dashed for i_h = i_housing.
	ax = axes[1]

	# Steady-state prices the policy was solved with
	r_ss = float(getattr(self.ss, 'r', getattr(par, 'r', 0.0)))
	w_ss = float(getattr(self.ss, 'w', 1.0))
	R    = 1.0 + r_ss
	q_u_ss = float(getattr(self.ss, 'q_u', 1.0))
	q_r_ss = float(getattr(self.ss, 'q_r', 1.0))
	f_u_ss = float(getattr(self.ss, 'f_u',
	                       q_u_ss * (R - (1.0 - par.theta) * (1.0 - par.delta_H))))
	f_r_ss = float(getattr(self.ss, 'f_r',
	                       q_r_ss * (R - (1.0 - par.theta) * (1.0 - par.delta_H))))
	h_l    = float(getattr(par, 'h_l', par.h_r))

	def _housing_cost_stay(i_h):
		"""Per-period housing cost when staying in current state (i_h_next = i_h)."""
		if i_h == 0:
			return f_r_ss * h_l
		if i_h == 1:
			return f_u_ss * h_l
		if i_h == 2:
			return par.lambda_ltv * q_r_ss * par.h_r * (r_ss + 1.0 / par.T_mort)
		return par.lambda_ltv * q_u_ss * par.h_u * (r_ss + 1.0 / par.T_mort)

	# z used for income / PIH (same as the productivity index passed in)
	z_val = float(par.z_grid[i_z]) if hasattr(par, 'z_grid') else 0.0

	def _labor_income(t):
		"""w · exp(χ[t] + z), with chi clamped to the available range."""
		if hasattr(par, 'chi') and len(par.chi) > 0:
			chi = float(par.chi[min(t, len(par.chi) - 1)])
		else:
			chi = 0.0
		return w_ss * np.exp(chi + z_val)

	a_lag_grid = np.linspace(0.0, a_hi_buffer, 400)

	# Per-state buffer-stock curves
	for i_h in state_list:
		hc      = _housing_cost_stay(i_h)
		y_t     = _labor_income(age)
		m_t     = R * a_lag_grid + y_t - hc
		m_t     = np.maximum(m_t, par.cfloor)
		pol_h   = pol[i_h][age][i_z]
		c_t     = np.interp(m_t, pol_h['m'], pol_h['c'])
		a_t     = m_t - c_t
		delta_a = a_t - a_lag_grid
		ax.plot(a_lag_grid, delta_a,
		        color=state_color[i_h], ls=state_ls[i_h], lw=2,
		        label=state_label[i_h])

	# PIH benchmark for the reference state i_housing (deterministic z, finite horizon)
	T_left = max(1, J_life - age)
	rho    = float(par.rho)
	beta   = float(par.beta)
	s_arr  = np.arange(T_left)
	# Annuity factor: A_t = Σ_{s=0}^{T_left-1} (β R)^{s/ρ} · R^{-s}
	A_t = float(np.sum((beta * R) ** (s_arr / rho) * R ** (-s_arr)))
	hc_ref = _housing_cost_stay(i_housing)
	# PV of future disposable income (s=1,...,T_left-1) at constant state i_housing
	if T_left > 1:
		s2  = np.arange(1, T_left)
		y_s = np.array([_labor_income(age + int(s)) for s in s2])
		H_t = float(np.sum((y_s - hc_ref) * R ** (-s2)))
	else:
		H_t = 0.0
	y_ref     = _labor_income(age)
	m_t_ref   = np.maximum(R * a_lag_grid + y_ref - hc_ref, par.cfloor)
	c_pih     = (m_t_ref + H_t) / A_t
	delta_pih = m_t_ref - c_pih - a_lag_grid
	ax.plot(a_lag_grid, delta_pih, 'k--', lw=1.5, alpha=0.7,
	        label=f'PIH ({curr_labels[i_housing]})')

	ax.axhline(0.0, color='0.4', lw=0.8, alpha=0.7)
	ax.set_xlim(0.0, a_hi_buffer)
	ax.set_xlabel(r'$a_{t-1}$  (savings carried in)')
	ax.set_ylabel(r'$a_t - a_{t-1}$')
	ax.set_title('Buffer-stock change in savings')
	ax.legend(fontsize=8)
	ax.grid(True, alpha=0.4)

	# ── Panel (2): Choice probabilities given current i_housing ─────────────
	ax = axes[2]
	mask_pr = m_grid <= m_hi_choice
	for k in range(4):
		prob = pr_reg[k, :]
		prob = np.where(np.isfinite(prob), prob, np.nan)
		ax.plot(m_grid[mask_pr], prob[mask_pr],
		        color=state_color[k], ls=state_ls[k], lw=2,
		        label=f'→ {next_labels[k]}')
	ax.set_xlim(m_lo, m_hi_choice)
	ax.set_ylim(0, 1)
	ax.set_xlabel('Cash-on-hand (m)')
	ax.set_ylabel('Choice probability')
	ax.set_title(f'Choice probs  |  current: {curr_labels[i_housing]}')
	ax.legend(fontsize=9)
	ax.grid(True, alpha=0.4)

	# ── Suptitle: age, fixed type and productivity level ────────────────────
	real_age = age_start + age
	fig.suptitle(
	    f'Policy functions  |  age={age} (real ~{real_age}),  '
	    f'fix={i_fix},  z-index={i_z}  (z={z_val:.3f})',
	    fontsize=11, y=1.02)

	plt.tight_layout()
	plt.show()


def plot_choice_probs(self, age=20, i_fix=0, i_z=0, m_max_plot=None, age_start=16):
	"""Plot all 4 next-period housing choice probabilities at a chosen age.

	Reads age-specific probabilities from ``self.hh_policy[i_h][age][i_z]['pr_choices']``
	(shape ``(4, N_endog)``) and interpolates them onto ``par.m_grid``.
	``self.sol.pr_choices`` is not used because the backward DC-EGM sweep
	overwrites it and only the age-0 policy survives — flat in m, not useful.

	Each panel corresponds to one current housing state and shows
	P(choose h_next | m, current_state, age) for all four choices h_next.

	Parameters
	----------
	age        : lifecycle index ``it`` (0 = newborn, J-1 = terminal)
	i_fix      : fixed-type index (kept for API compatibility)
	i_z        : productivity index
	m_max_plot : x-axis upper limit (default: full grid)
	age_start  : real age corresponding to model age 0 (default 16)
	"""
	par = self.par
	_ = i_fix  # currently unused; kept for API compatibility

	if not hasattr(self, 'hh_policy'):
		raise AttributeError(
			"self.hh_policy not available — run solve_hh_backwards(model) first."
		)

	J_life = int(getattr(par, 'J', getattr(par, 'T', 0)))
	age = int(age)
	if age < 0 or age >= J_life:
		raise ValueError(f"age must be in [0, {J_life - 1}], got {age}")

	pol = self.hh_policy
	for i_h in range(4):
		if age not in pol.get(i_h, {}) or i_z not in pol[i_h][age]:
			raise KeyError(
				f"hh_policy missing entry for i_h={i_h}, age={age}, i_z={i_z}"
			)

	m_grid = par.m_grid if hasattr(par, 'm_grid') else par.a_grid
	m_hi = float(np.nanmax(m_grid)) if np.any(np.isfinite(m_grid)) else 1.0
	if m_max_plot is not None:
		m_hi = min(float(m_max_plot), m_hi)
	mask = m_grid <= m_hi

	# Interpolate pr_choices (4, N_endog) onto m_grid for each current housing state.
	pr_by_curr = {}
	for i_h_curr in range(4):
		pol_h = pol[i_h_curr][age][i_z]
		m_endog = pol_h['m']
		pr_endog = pol_h['pr_choices']  # (4, N_endog)
		pr_reg = np.empty((4, m_grid.size))
		for k in range(4):
			pr_reg[k, :] = np.interp(m_grid, m_endog, pr_endog[k, :])
		pr_reg = np.clip(pr_reg, 0.0, 1.0)
		col_sums = pr_reg.sum(axis=0, keepdims=True)
		col_sums = np.where(col_sums > 0.0, col_sums, 1.0)
		pr_by_curr[i_h_curr] = pr_reg / col_sums

	curr_labels = {
		0: 'Current: rural renter',
		1: 'Current: urban renter',
		2: 'Current: rural owner',
		3: 'Current: urban owner',
	}
	next_labels = {
		0: 'rural renter',
		1: 'urban renter',
		2: 'rural owner',
		3: 'urban owner',
	}
	next_colors = {0: 'tab:blue', 1: 'tab:orange', 2: 'tab:green', 3: 'tab:red'}
	next_ls     = {0: '-', 1: '-', 2: '--', 3: '--'}

	fig, axes = plt.subplots(2, 2, figsize=(12, 8), sharey=True)

	for i_h_curr, ax in zip(range(4), axes.flat):
		pr_reg = pr_by_curr[i_h_curr]
		for k in range(4):
			prob = pr_reg[k, :]
			prob = np.where(np.isfinite(prob), prob, np.nan)
			ax.plot(m_grid[mask], prob[mask],
			        color=next_colors[k], ls=next_ls[k], lw=2,
			        label=f'→ {next_labels[k]}')
		ax.set_title(curr_labels[i_h_curr], fontsize=10)
		ax.set_xlabel('Cash-on-hand (m)', fontsize=9)
		ax.set_ylabel('Choice probability', fontsize=9)
		ax.set_xlim(0, m_hi)
		ax.set_ylim(0, 1.0)
		ax.legend(fontsize=8)
		ax.grid(True, alpha=0.4)

	z_val = float(par.z_grid[i_z]) if hasattr(par, 'z_grid') else i_z
	real_age = age_start + age
	fig.suptitle(
		f'Next-period housing choice probabilities  |  age={age} (real ~{real_age}),  '
		f'fix={i_fix},  z-index={i_z}  (z={z_val:.3f})',
		fontsize=11, y=1.01)
	plt.tight_layout()
	plt.show()


def plot_choice_probs_vs_savings(self, i_h_current=0, i_z_next=0, t=None, a_max_plot=None):
	"""Plot z'-conditional choice probabilities directly against the SAVINGS grid.

	This bypasses the EGM change-of-variables m=a+c(a) and shows the raw softmax
	probabilities as a function of savings a — these are smooth by construction.
	Useful to confirm that kinkiness in plot_choice_probs comes from the EGM
	change-of-variables, not from the probability computation itself.

	Reads from model.hh_policy (endogenous grid) — solve household problem first.

	Parameters
	----------
	i_h_current : current housing state (0–3)
	i_z_next    : which realised z' column of pr_choices_asav to plot (0..Nz-1)
	t           : period index (default: first non-terminal period, T-2)
	a_max_plot  : x-axis limit on savings a (default: full savings grid)
	"""
	par = self.par
	sol = self.sol

	if not hasattr(sol, 'pr_choices_asav'):
		print("sol.pr_choices_asav not available — solve the household problem first.")
		return

	a_grid = par.a_grid
	a_hi = float(np.nanmax(a_grid))
	if a_max_plot is not None:
		a_hi = min(float(a_max_plot), a_hi)
	mask = a_grid <= a_hi

	# pr_choices_asav: (Nfix, Nz_next, Nh_curr, 4, Na)
	# index [0, i_z_next, i_h_current, k, :] gives P(h'=k | a, z'=i_z_next, h_curr=i_h_current)
	next_labels = {0: 'rural renter', 1: 'urban renter', 2: 'rural owner', 3: 'urban owner'}
	next_colors = {0: 'tab:blue', 1: 'tab:orange', 2: 'tab:green', 3: 'tab:red'}
	next_ls     = {0: '-', 1: '-', 2: '--', 3: '--'}
	curr_labels = {0: 'rural renter', 1: 'urban renter', 2: 'rural owner', 3: 'urban owner'}

	fig, ax = plt.subplots(figsize=(9, 5))
	for k in range(4):
		prob = sol.pr_choices_asav[0, i_z_next, i_h_current, k, :]
		prob = np.where(np.isfinite(prob), prob, np.nan)
		ax.plot(a_grid[mask], prob[mask],
		        color=next_colors[k], ls=next_ls[k], lw=2,
		        label=f'→ {next_labels[k]}')

	z_next_val = float(par.z_grid[i_z_next]) if hasattr(par, 'z_grid') else i_z_next
	ax.set_title(
		f"P(h' | savings a, z'={z_next_val:.3f})  —  current: {curr_labels[i_h_current]}\n"
		f"(x-axis = savings a, NOT cash-on-hand m; smooth by construction)",
		fontsize=10)
	ax.set_xlabel("Next-period savings (a')")
	ax.set_ylabel("Choice probability")
	ax.set_xlim(0, a_hi)
	ax.set_ylim(0, 1)
	ax.legend(fontsize=9)
	ax.grid(True, alpha=0.4)
	plt.tight_layout()
	plt.show()


def plot_value_function(self, t, i_fix=0, i_z=0, idc_prev=0, include_logsum=True, m_max=100):
	"""Plot value functions for all 4 housing states at period t.

	Parameters
	----------
	t              : period index (0-based)
	i_fix          : fixed-type index (unused in DC-EGM core)
	i_z            : productivity index (unused; kept for compatibility)
	idc_prev       : unused, kept for compatibility
	include_logsum : overlay the inclusive (ex-ante) value from the logsum
	m_max          : x-axis upper limit (default 100)
	"""
	par = self.par

	if not hasattr(self, 'hh_value'):
		print("Full household solution not available. Run solve_hh_backwards() first.")
		return

	value = self.hh_value
	T = par.T if hasattr(par, 'T') else par.Tbar
	if t < 0 or t >= T:
		print(f"t must be in [0, {T-1}]")
		return

	# Check all 4 states are available
	missing = [h for h in range(4) if t not in value.get(h, {})]
	if missing:
		print(f"Value functions not available for t={t}, states={missing}")
		return

	state_label = {
		0: 'V(m | rural renter)',
		1: 'V(m | urban renter)',
		2: 'V(m | rural owner)',
		3: 'V(m | urban owner)',
	}
	state_color = {0: 'tab:blue', 1: 'tab:orange', 2: 'tab:green', 3: 'tab:red'}
	state_ls    = {0: '--', 1: '--', 2: '-', 3: '-'}

	m_plot = np.linspace(0.0, min(float(par.mmax), float(m_max)), 600)

	v_by_state = {}
	for h in range(4):
		vf = value[h][t][i_z]
		v_by_state[h] = np.interp(m_plot, vf['m'], vf['v'])

	fig, ax = plt.subplots(figsize=(12, 6))

	for h in range(4):
		ax.plot(m_plot, v_by_state[h],
		        color=state_color[h], ls=state_ls[h], lw=2.2, alpha=0.9,
		        label=state_label[h])

	if include_logsum:
		from household_problem import HousingModel
		hh   = HousingModel(self)
		v_stack     = np.vstack([v_by_state[h] for h in range(4)])  # (4, N)
		v_inclusive = hh._logsum(v_stack)
		ax.plot(m_plot, v_inclusive, color='black', lw=2.5, ls='-', alpha=0.85,
		        label='Inclusive value (logsum over 4 states)')

	ax.set_xlabel('Cash-on-hand (m)', fontsize=12)
	ax.set_ylabel('Value', fontsize=12)
	ax.set_xlim(0.0, m_plot[-1])
	ax.set_title(f'Value Functions at t={t}  (dashed = renters, solid = owners)', fontsize=13)
	ax.legend(fontsize=10, loc='best')
	ax.grid(True, alpha=0.3)
	plt.tight_layout()
	plt.show()


def plot_lifecycle_wealth(model, age_start=16, title='Life-cycle Wealth Profiles'):
	"""Plot mean liquid wealth and housing wealth over the life cycle.

	Three panels side-by-side: Aggregate, Rural (states 0,2), Urban (states 1,3).
	Requires ss.D_cohort to be available.

	Parameters:
	    model:      model instance with ss, sol, par populated
	    age_start:  real age corresponding to model age 0 (default 25)
	    title:      figure suptitle
	"""
	ss  = model.ss
	par = model.par
	sol = model.sol

	if not hasattr(ss, 'D_cohort') or ss.D_cohort is None:
		raise AttributeError(
			'ss.D_cohort not found — solve the steady state with demographics enabled first.')

	D_cohort = ss.D_cohort  # (J, Nz, Nh, Na)
	J        = D_cohort.shape[0]
	ages     = np.arange(age_start, age_start + J)

	a_pol = sol.a[0]  # (Nz, Nh, Na) — savings policy over m_grid

	# Housing asset value per state: renters own nothing, owners hold q*h
	h_wealth = np.array([
		0.0,                    # 0 = rural_renter
		0.0,                    # 1 = urban_renter
		par.h_r * ss.q_r,       # 2 = rural_owner
		par.h_u * ss.q_u,       # 3 = urban_owner
	])

	def _moments(D_j, h_idx):
		"""Mean liquid and housing wealth per capita for given housing-state indices."""
		D_sub = D_j[:, h_idx, :]               # (Nz, |h_idx|, Na)
		mass  = D_sub.sum()
		if mass < 1e-15:
			return np.nan, np.nan
		liquid  = np.sum(D_sub * a_pol[:, h_idx, :]) / mass
		housing = np.sum(D_sub * h_wealth[h_idx][np.newaxis, :, np.newaxis]) / mass
		return liquid, housing

	all_idx   = np.array([0, 1, 2, 3])
	rural_idx = np.array([0, 2])
	urban_idx = np.array([1, 3])

	liq_agg   = np.zeros(J); hw_agg   = np.zeros(J)
	liq_rural = np.zeros(J); hw_rural = np.zeros(J)
	liq_urban = np.zeros(J); hw_urban = np.zeros(J)

	for j in range(J):
		D_j = D_cohort[j]
		liq_agg[j],   hw_agg[j]   = _moments(D_j, all_idx)
		liq_rural[j], hw_rural[j] = _moments(D_j, rural_idx)
		liq_urban[j], hw_urban[j] = _moments(D_j, urban_idx)

	fig, axes = plt.subplots(1, 3, figsize=(15, 4.5))

	panels = [
		(axes[0], 'Aggregate', liq_agg,   hw_agg),
		(axes[1], 'Rural',     liq_rural, hw_rural),
		(axes[2], 'Urban',     liq_urban, hw_urban),
	]

	for ax, label, liq, hw in panels:
		ax.plot(ages, liq, lw=2, color='steelblue',  label='Liquid wealth')
		ax.plot(ages, hw,  lw=2, color='firebrick', linestyle='--', label='Housing wealth')
		ax.axhline(0, color='k', lw=0.6)
		ax.set_title(label, fontsize=13)
		ax.set_xlabel('Age')
		ax.set_ylabel('Mean wealth per capita')
		ax.legend(fontsize=10)
		ax.grid(True, alpha=0.3)

	fig.suptitle(title, fontsize=14)
	fig.tight_layout()
	plt.show()
	plt.close(fig)


def plot_lifecycle_tenure(model, age_start=16, title='Life-cycle Tenure Profiles'):
	"""Plot housing tenure shares over the life cycle.

	Three panels side-by-side: Aggregate, Rural (conditional on being rural),
	Urban (conditional on being urban). Each panel shows stacked areas for
	owner and renter shares.

	Requires ss.D_cohort.

	Parameters:
	    model:      model instance with ss.D_cohort, par populated
	    age_start:  real age at model age 0 (default 16)
	    title:      figure suptitle
	"""
	ss  = model.ss

	if not hasattr(ss, 'D_cohort') or ss.D_cohort is None:
		raise AttributeError(
			'ss.D_cohort not found — solve the steady state with demographics enabled first.')

	D_cohort = ss.D_cohort  # (J, Nz, Nh, Na)
	J        = D_cohort.shape[0]
	ages     = np.arange(age_start, age_start + J)

	# State indices: 0=rural_renter, 1=urban_renter, 2=rural_owner, 3=urban_owner
	# Aggregate
	share_owner       = np.zeros(J)
	share_renter      = np.zeros(J)
	# Rural conditional: P(rural_owner | rural), P(rural_renter | rural)
	own_rural         = np.zeros(J)
	rent_rural        = np.zeros(J)
	# Urban conditional: P(urban_owner | urban), P(urban_renter | urban)
	own_urban         = np.zeros(J)
	rent_urban        = np.zeros(J)

	for j in range(J):
		D_j   = D_cohort[j]  # (Nz, Nh, Na)
		total = D_j.sum()
		if total < 1e-15:
			share_owner[j] = share_renter[j] = np.nan
			own_rural[j] = rent_rural[j] = own_urban[j] = rent_urban[j] = np.nan
			continue

		share_owner[j]  = sum(D_j[:, h, :].sum() for h in [2, 3]) / total
		share_renter[j] = sum(D_j[:, h, :].sum() for h in [0, 1]) / total

		rural_total = sum(D_j[:, h, :].sum() for h in [0, 2])
		urban_total = sum(D_j[:, h, :].sum() for h in [1, 3])

		if rural_total > 1e-15:
			own_rural[j]  = D_j[:, 2, :].sum() / rural_total
			rent_rural[j] = D_j[:, 0, :].sum() / rural_total
		else:
			own_rural[j] = rent_rural[j] = np.nan

		if urban_total > 1e-15:
			own_urban[j]  = D_j[:, 3, :].sum() / urban_total
			rent_urban[j] = D_j[:, 1, :].sum() / urban_total
		else:
			own_urban[j] = rent_urban[j] = np.nan

	fig, axes = plt.subplots(1, 3, figsize=(15, 4.5))

	panels = [
		(axes[0], 'Aggregate',       share_owner,  share_renter),
		(axes[1], 'Rural (conditional)', own_rural, rent_rural),
		(axes[2], 'Urban (conditional)', own_urban, rent_urban),
	]

	for ax, label, own, rent in panels:
		ax.stackplot(ages, own, rent,
		             labels=['Owner', 'Renter'],
		             colors=['firebrick', 'steelblue'], alpha=0.75)
		ax.set_ylim(0, 1)
		ax.set_title(label, fontsize=13)
		ax.set_xlabel('Age')
		ax.set_ylabel('Population share')
		ax.legend(fontsize=10, loc='upper left')
		ax.grid(True, alpha=0.3)

	fig.suptitle(title, fontsize=14)
	fig.tight_layout()
	plt.show()
	plt.close(fig)


def _load_makro_lifecycle(model=None, fname='MAKRO_life_cycle_profiles.xlsx'):
	"""Load the MAKRO age-profile xlsx and return a dict {column_name: (ages, values)}.

	The file is expected to have an 'Age' column plus one or more named series
	(currently: 'vW', 'vHhRealKred', 'vBolig', 'vHhx_scaled'). Looked up next to
	HANCHousingModel.py when `model` is given, otherwise next to this plots.py.
	"""
	import os as _os
	import openpyxl as _openpyxl
	if model is not None:
		# co-located with HANCHousingModel.py
		import HANCHousingModel as _hm
		base = _os.path.dirname(_os.path.abspath(_hm.__file__))
	else:
		base = _os.path.dirname(_os.path.abspath(__file__))
	path = _os.path.join(base, fname)
	wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)
	ws = wb.active
	rows = list(ws.iter_rows(values_only=True))
	wb.close()
	header = [str(c).strip() if c is not None else '' for c in rows[0]]
	i_age = header.index('Age')
	ages = np.array([int(r[i_age]) for r in rows[1:] if r[i_age] is not None])
	out = {}
	for j, col in enumerate(header):
		if not col or col == 'Age':
			continue
		vals = np.array([
			float(r[j]) if (r[j] is not None and str(r[j]).strip() != '') else np.nan
			for r in rows[1:] if r[i_age] is not None
		], dtype=float)
		out[col] = (ages, vals)
	return out


def plot_lifecycle_wealth_pe(pe, model, year=2024, year0=1992, age_start=16,
                             title=None, q_u_path=None, q_r_path=None,
                             scenario=None, data_overlay=True, data_anchor_age=75):
	"""PE analogue of plot_lifecycle_wealth, sliced at a chosen calendar year.

	Three panels side-by-side: Aggregate, Rural (states 0,2), Urban (states 1,3).
	Reads the age x housing-state x net-wealth-bin histogram stored on the
	PE result (`pe['nw_hist']`) and splits mean net wealth into liquid and
	gross-housing components using the housing-price path and the calibrated
	housing-service levels `model.par.h_*`.

	Price-path source — first match wins:
	    1. Explicit `q_u_path`, `q_r_path` keyword args (1-D arrays of length T).
	    2. `scenario` dict containing `'q_u'` and `'q_r'` (e.g. baseline_scenario).
	    3. The PE result's own `pe['q_u_path']`, `pe['q_r_path']` (default).
	The override lets you bypass a buggy `pe['q_u_path']` when you know the
	scenario dict actually fed into the simulation has the right values.

	Args:
	    pe              : pe_result dict from simulate_olg_pe.
	    model           : HANCHousingModel — needed for par.h_u, par.h_r, par.lambda_ltv.
	    year            : calendar year at which to slice the lifecycle (default 2024).
	    year0           : calendar year of t=0 (default 1992).
	    age_start       : real age corresponding to model age 0 (default 16).
	    title           : figure suptitle (default auto-generated).
	    q_u_path        : optional override for the urban price path (length T).
	    q_r_path        : optional override for the rural price path (length T).
	    scenario        : optional scenario dict from which 'q_u' and 'q_r' are
	                      read if q_u_path/q_r_path not given.
	    data_overlay    : if True, overlay MAKRO_life_cycle_profiles.xlsx on the
	                      Aggregate panel: net housing wealth = vBolig − vHhRealKred
	                      and liquid wealth = vHhx_scaled. The data series are
	                      rescaled so that the data's net housing wealth at
	                      `data_anchor_age` equals the model's housing wealth at
	                      the oldest model age (J−1). The same scale is applied
	                      to the liquid series.
	    data_anchor_age : age (in years) at which model and data housing wealth
	                      are matched for scaling (default 77 — picked to avoid
	                      the model-side dip around age 80).
	"""
	par = model.par
	nw_hist = np.asarray(pe['nw_hist'], dtype=float)   # (T, J, Nh, n_bins)
	nw_grid = np.asarray(pe['nw_grid'], dtype=float)   # (n_bins,)
	T, J, Nh, _ = nw_hist.shape

	def _resolve_q(name):
		# 1. explicit override
		explicit = q_u_path if name == 'q_u' else q_r_path
		if explicit is not None:
			return np.asarray(explicit, dtype=float).ravel()
		# 2. scenario dict
		if scenario is not None and name in scenario:
			return np.asarray(scenario[name], dtype=float).ravel()
		# 3. PE result fallback
		return np.asarray(pe[name + '_path'], dtype=float).ravel()

	q_u_arr = _resolve_q('q_u')
	q_r_arr = _resolve_q('q_r')

	t = int(year - year0)
	if not (0 <= t < T):
		raise ValueError(f'year={year} outside the PE horizon [{year0}, {year0+T-1}].')

	q_u_t = float(q_u_arr[t])
	q_r_t = float(q_r_arr[t])
	h_u   = float(par.h_u)
	h_r   = float(par.h_r)
	one_minus_ltv = 1.0 - float(par.lambda_ltv)

	# Per state at year t: gross housing wealth (owner only) and equity (owner only).
	# State indices: 0=rural_renter, 1=urban_renter, 2=rural_owner, 3=urban_owner.
	gross_h = np.array([0.0, 0.0, q_r_t * h_r, q_u_t * h_u])              # (Nh,)
	equity  = one_minus_ltv * gross_h                                       # (Nh,)

	# Per-(age, state) mass and mean net wealth at year t.
	mass_age_h = nw_hist[t].sum(axis=-1)                                    # (J, Nh)
	nw_mean    = np.where(mass_age_h > 1e-15,
	                      (nw_hist[t] * nw_grid[None, None, :]).sum(axis=-1) /
	                      np.maximum(mass_age_h, 1e-15),
	                      0.0)                                              # (J, Nh)
	# Mean liquid wealth in (age, state) = mean net wealth minus state-constant equity.
	liquid_age_h  = nw_mean - equity[None, :]                               # (J, Nh)
	housing_age_h = np.broadcast_to(gross_h[None, :], (J, Nh))              # (J, Nh)

	def _agg(idx):
		"""Mass-weighted mean liquid and housing wealth over state subset `idx`."""
		mass = mass_age_h[:, idx]                                            # (J, k)
		total = mass.sum(axis=-1)                                            # (J,)
		safe  = np.maximum(total, 1e-15)
		liq = (liquid_age_h[:, idx]  * mass).sum(axis=-1) / safe             # (J,)
		hw  = (housing_age_h[:, idx] * mass).sum(axis=-1) / safe             # (J,)
		liq = np.where(total > 1e-15, liq, np.nan)
		hw  = np.where(total > 1e-15, hw,  np.nan)
		return liq, hw

	all_idx   = np.array([0, 1, 2, 3])
	rural_idx = np.array([0, 2])
	urban_idx = np.array([1, 3])

	liq_agg,   hw_agg   = _agg(all_idx)
	liq_rural, hw_rural = _agg(rural_idx)
	liq_urban, hw_urban = _agg(urban_idx)

	ages = np.arange(age_start, age_start + J)

	# Prepare MAKRO data overlay for the Aggregate panel.
	# Anchor at `data_anchor_age` in BOTH series so the data and model housing-
	# wealth lines visually intersect at that age. If the model's lifecycle does
	# not reach `data_anchor_age`, fall back to the model's oldest age and tell
	# the user where the anchor actually lives.
	data_ages = data_hw_scaled = data_liq_scaled = None
	anchor_age_used = None
	if data_overlay:
		try:
			makro = _load_makro_lifecycle(model=model)
			d_ages,   v_bolig    = makro['vBolig']
			_,        v_kred     = makro['vHhRealKred']
			_,        v_liq      = makro['vHhx_scaled']
			d_hw_net  = v_bolig - v_kred                # net housing wealth

			model_last_age = age_start + len(hw_agg) - 1
			anchor_age_used = int(min(int(data_anchor_age), model_last_age))
			if anchor_age_used != int(data_anchor_age):
				print(f"  [plot_lifecycle_wealth_pe] data_anchor_age={data_anchor_age} "
				      f"is past the model's last age ({model_last_age}); "
				      f"anchoring at age {anchor_age_used} instead.")

			# Model housing wealth at the anchor age
			model_idx = anchor_age_used - age_start
			hw_model_at_anchor = float(hw_agg[model_idx])

			# Data housing wealth at the anchor age
			data_anchor_idx = np.where(d_ages == anchor_age_used)[0]
			if data_anchor_idx.size == 0:
				raise ValueError(
					f"anchor age {anchor_age_used} not present in MAKRO ages "
					f"{d_ages.min()}–{d_ages.max()}")
			d_anchor = float(d_hw_net[data_anchor_idx[0]])

			if not (np.isfinite(hw_model_at_anchor) and np.isfinite(d_anchor) and d_anchor > 0):
				raise ValueError(
					f"cannot scale data overlay (hw_model[age={anchor_age_used}]"
					f"={hw_model_at_anchor}, data_hw[{anchor_age_used}]={d_anchor})")

			scale = hw_model_at_anchor / d_anchor
			data_ages       = d_ages
			data_hw_scaled  = scale * d_hw_net
			data_liq_scaled = scale * v_liq
		except Exception as e:
			print(f"  [plot_lifecycle_wealth_pe] data overlay disabled: {e}")
			data_ages = data_hw_scaled = data_liq_scaled = None
			anchor_age_used = None

	fig, axes = plt.subplots(1, 3, figsize=(10, 7))
	panels = [
		(axes[0], 'Aggregate', liq_agg,   hw_agg),
		(axes[1], 'Rural region',     liq_rural, hw_rural),
		(axes[2], 'Urban region',     liq_urban, hw_urban),
	]

	col_house = '#C92828'   # red
	col_liquid = '#003C8F'   # dark blue

	for ax, label, liq, hw in panels:
		ax.plot(ages, liq, lw=2, color=col_liquid,  label='Liquid wealth')
		ax.plot(ages, hw,  lw=2, color=col_house,  label='Housing wealth')

		# MAKRO data overlay only on the Aggregate panel.
		if label == 'Aggregate' and data_ages is not None:
			ax.plot(data_ages, data_hw_scaled, lw=1.5, color=col_house,
			        linestyle='--', alpha=0.85,
			        label=f'Data')
			ax.plot(data_ages, data_liq_scaled, lw=1.5, color=col_liquid,
			        linestyle='--', alpha=0.85, label='Data')

		ax.axhline(0, color='k', lw=0.6)
		ax.set_title(label, fontsize=13)
		ax.set_xlim(16, 80)
		ax.set_xlabel('Age')
		ax.set_ylabel('Mean wealth per capita')
		ax.legend(fontsize=8 if (label == 'Aggregate' and data_ages is not None) else 10, frameon=False)
		ax.grid(False)
		ax.set_box_aspect(1.0)


	if title is None:
		title = f'Life-cycle wealth profiles — PE, year {year}'
	fig.suptitle(title, fontsize=14)
	fig.tight_layout()
	plt.show()
	plt.close(fig)


def plot_makro_lifecycle_profiles(model=None, title=None,
                                  fname='MAKRO_life_cycle_profiles.xlsx'):
	"""Plot the raw MAKRO age profiles in a 1x2 figure, in the graphical style
	of plot_lifecycle_wealth_pe.

	Left panel  : earnings (vW).
	Right panel : vHhRealKred (mortgage debt), vBolig (property value) and
	              vHhx_scaled (financial wealth).

	`model` only controls which directory the xlsx is looked up in (passed to
	`_load_makro_lifecycle`); the data itself is model-independent, so it may be
	left as None.
	"""
	makro = _load_makro_lifecycle(model=model, fname=fname)

	col_house  = '#C92828'   # red
	col_liquid = '#003C8F'   # dark blue
	col_debt   = '#E08214'   # orange

	fig, axes = plt.subplots(1, 2, figsize=(9, 4.6))

	# --- Left: earnings (vW) ------------------------------------------------
	ax = axes[0]
	if 'vW' in makro:
		ages, vW = makro['vW']
		ax.plot(ages, vW, lw=2, color=col_liquid, label='vW (earnings)')
	ax.axhline(0, color='k', lw=0.6)
	ax.set_title('Earnings', fontsize=13)
	ax.set_xlabel('Age')
	ax.set_ylabel('Mean earnings per capita')
	ax.legend(fontsize=10, frameon=False)
	ax.grid(False)
	ax.set_box_aspect(1.0)

	# --- Right: balance-sheet components ------------------------------------
	ax = axes[1]
	right_series = [
		('vHhRealKred', 'vHhRealKred (mortgage debt)',   col_debt),
		('vBolig',      'vBolig (property value)',        col_house),
		('vHhx_scaled', 'vHhx_scaled (financial wealth)', col_liquid),
	]
	for col, lbl, color in right_series:
		if col in makro:
			ages, vals = makro[col]
			ax.plot(ages, vals, lw=2, color=color, label=lbl)
	ax.axhline(0, color='k', lw=0.6)
	ax.set_title('Wealth components', fontsize=13)
	ax.set_xlabel('Age')
	ax.set_ylabel('Mean wealth per capita')
	ax.legend(fontsize=9, frameon=False)
	ax.grid(False)
	ax.set_box_aspect(1.0)

	# Match the x-range to the data span on both panels.
	_ages_all = [v[0] for v in makro.values() if len(v[0]) > 0]
	if _ages_all:
		amin = min(float(a.min()) for a in _ages_all)
		amax = max(float(a.max()) for a in _ages_all)
		for ax in axes:
			ax.set_xlim(amin, amax)

	if title is None:
		title = 'MAKRO life-cycle profiles'
	fig.suptitle(title, fontsize=14)
	fig.tight_layout()
	plt.show()
	plt.close(fig)
	return fig




def load_datagraphs(filepath, start_year=1992):
	"""Load empirical data from datagraphs.xlsx.

	Returns a dict mapping panel title (as used in plot_transition_paths) to
	a tuple (years, values) of numpy arrays.  Rows with all-NaN values are
	skipped.  The year range is inferred from the number of data rows.

	Parameters
	----------
	filepath   : str – path to the xlsx file
	start_year : int – calendar year of the first data row (default 1992)
	"""
	import openpyxl

	def _norm_header(x):
		# Robust to whitespace/case differences in Excel headers.
		return ' '.join(str(x).strip().split()).lower()

	# Mapping: xlsx column header → plot panel title
	_COL_TO_TITLE = {
		'Capital/GDP':            'Capital/GDP',
		'Urban housing/GDP':      'Urban housing/GDP',
		'Rural housing/GDP':      'Rural housing/GDP',
		'Urban housing wealth':   'Urban housing wealth',
		'Rural housing wealth':   'Rural housing wealth',
		'Top 1':                  'Top 1% wealth share',
		'Top 10':                 'Top 10% wealth share',
		'Middle 40':              'Middle 40% wealth share',
		'Bottom 50':              'Bottom 50% wealth share',
		'Urban price/rural price':'Urban/rural price ratio',
		'Urban population share': 'Urban population share',
		'Urban housing density':  'Urban housing density',
		'Rural housing density':  'Rural housing density',
		'Investment/GDP':         'Investment/GDP',
		'Urban construction/GDP': 'Urban construction/GDP',
		'Rural construction/GDP': 'Rural construction/GDP',
		'Real rate':			  'Real rate',
		'Net Foreign Assets':     'Net Foreign Assets/GDP',
		'Net Foreign Assets/GDP': 'Net Foreign Assets/GDP',
		'Share of renters - Urban Region': 'Share of renters - Urban region',
		'Share of renters - Urban region': 'Share of renters - Urban region',
		'Share of renters - Rural Region': 'Share of renters - Rural region',
		'Share of renters - Rural region': 'Share of renters - Rural region',
	}
	_COL_TO_TITLE_NORM = {_norm_header(k): v for k, v in _COL_TO_TITLE.items()}

	wb  = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
	ws  = wb.active
	rows = list(ws.iter_rows(values_only=True))
	wb.close()

	header   = [str(c).strip() if c is not None else '' for c in rows[0]]
	data_rows = rows[1:]
	n_rows    = len(data_rows)
	years     = np.arange(start_year, start_year + n_rows)

	result = {}
	for col_idx, col_name in enumerate(header):
		title = _COL_TO_TITLE_NORM.get(_norm_header(col_name))
		if title is None:
			continue
		vals = np.array(
			[float(r[col_idx]) if (r[col_idx] is not None and
			                        str(r[col_idx]).strip() != '') else np.nan
			 for r in data_rows],
			dtype=float,
		)
		if np.any(np.isfinite(vals)):
			result[title] = (years, vals)

	return result


def plot_transition_paths(self, T_plot=None, include_clearing=True, start_year=1992,
                          data_overlay=None):
	"""Plot transition paths for the moments listed in Table X of the paper.

	Groups:
	  Wealth ratios      : Capital/GDP, Urban housing/GDP, Rural housing/GDP
	  Wealth inequality  : Top 1%, Top 10%, Bottom 50% (asset share)
	  Housing            : Urban/rural price ratio, Urban pop share,
	                       Urban housing density, Rural housing density
	  Production         : Investment/GDP, Urban construction/GDP, Rural construction/GDP
	  Prices             : Real wage, Real interest rate, Urban house price, Rural house price

	Parameters
	----------
	T_plot        : int, optional – number of periods to plot (default: full horizon)
	include_clearing : bool – if True, append a market-clearing residual panel
	start_year    : int – calendar year of t=0 for x-axis labels (default 1992)
	data_overlay  : dict or str, optional – if a str, treated as a path to an xlsx
	                file loaded via load_datagraphs(); if a dict (from load_datagraphs),
	                used directly. Maps panel title → (years, values) and overlays
	                the empirical series as dots on each matching panel.
	"""

	if not hasattr(self, 'path'):
		print("Path object not available on model.")
		return

	# Resolve data_overlay
	if isinstance(data_overlay, str):
		data_overlay = load_datagraphs(data_overlay, start_year=start_year)

	par = self.par
	ss  = self.ss
	path = self.path

	T = par.T if hasattr(par, 'T') else 1
	if T_plot is None:
		T_plot = T
	T_plot = int(max(1, min(T_plot, T)))
	t_grid = np.arange(T_plot) + start_year

	def _get(name):
		if not hasattr(path, name):
			return None
		arr = np.asarray(getattr(path, name))
		return arr[:T_plot, 0] if arr.ndim == 2 else arr[:T_plot]

	def _ss(name):
		return float(getattr(ss, name)) if hasattr(ss, name) else np.nan

	# ------------------------------------------------------------------
	# Build derived series
	# ------------------------------------------------------------------
	Y    = _get('Y')
	K    = _get('K')
	q_u  = _get('q_u')
	q_r  = _get('q_r')
	H_u  = _get('H_u')
	H_r  = _get('H_r')
	IH_u = _get('IH_u')
	IH_r = _get('IH_r')
	I    = _get('I')
	w    = _get('w')
	r    = _get('r')
	A_hh = _get('A_hh')

	#GDP definition as per equation (3.21):
	Y_tot = Y + q_u * IH_u + q_r * IH_r

	# Safe denominator
	Y_safe = np.where((Y_tot is not None) & np.isfinite(Y_tot) & (Y_tot > 0), Y_tot, np.nan) if Y_tot is not None else None

	def _ratio(num, den_safe, label):
		if num is None or den_safe is None:
			return None
		return np.where(np.isfinite(num) & np.isfinite(den_safe), num / den_safe, np.nan)

	# Wealth ratios
	K_Y    = _ratio(K,               Y_safe, 'K/Y')
	quHu_Y = _ratio(q_u * H_u if (q_u is not None and H_u is not None) else None, Y_safe, 'quHu/Y')
	qrHr_Y = _ratio(q_r * H_r if (q_r is not None and H_r is not None) else None, Y_safe, 'qrHr/Y')

	# Housing ratios
	price_ratio = (q_u / np.where(q_r > 0, q_r, np.nan)
	               if (q_u is not None and q_r is not None) else None)

	# Urban population share: convert household housing demand back to mass
	# of households via h_u, h_r so it matches the calibration moment
	# (D[:,1,:].sum() + D[:,3,:].sum()) / total_mass.
	H_u_hh = _get('H_u_hh')
	H_r_hh = _get('H_r_hh')
	if H_u_hh is not None and H_r_hh is not None:
		N_u = H_u_hh / par.h_u
		N_r = H_r_hh / par.h_r
		N_total = N_u + N_r
		urb_share = np.where(N_total > 0, N_u / N_total, np.nan)
	else:
		urb_share = None

	# Housing density: price * stock / GDP (value of housing per unit GDP)
	urb_density = _ratio(H_u if (q_u is not None and H_u is not None) else None,
	                     par.X_u, 'urb_density')
	rur_density = _ratio(H_r if (q_r is not None and H_r is not None) else None,
	                     par.X_r, 'rur_density')

	# Investment and construction ratios
	I_Y     = _ratio(I,    Y_safe, 'I/Y')
	quIHu_Y = _ratio(q_u * IH_u if (q_u is not None and IH_u is not None) else None, Y_safe, 'quIHu/Y')
	qrIHr_Y = _ratio(q_r * IH_r if (q_r is not None and IH_r is not None) else None, Y_safe, 'qrIHr/Y')

	# Wealth inequality: asset-share quantiles from simulated path.D
	# Populated by transition.compute_wealth_inequality(); skip gracefully if absent.
	top1_share  = _get('top1_share')
	top10_share = _get('top10_share')
	bot50_share = _get('bot50_share')

	# ------------------------------------------------------------------
	# Panel layout: 4 groups, each drawn in a separate figure row
	# ------------------------------------------------------------------
	specs = [
		# (series, title, y-label, ss_val_name)
		# --- Wealth ratios ---
		(K_Y,        'Capital/GDP',               'ratio', None),
		(quHu_Y,     'Urban housing/GDP',          'ratio', None),
		(qrHr_Y,     'Rural housing/GDP',          'ratio', None),
		# --- Wealth inequality ---
		(top1_share,  'Top 1% wealth share',       'share', None),
		(top10_share, 'Top 10% wealth share',      'share', None),
		(bot50_share, 'Bottom 50% wealth share',   'share', None),
		# --- Housing ---
		(price_ratio, 'Urban/rural price ratio',   'q_u/q_r', None),
		(urb_share,   'Urban population share',    'share',   None),
		(urb_density, 'Urban housing density',     'q_u·H_u/Y', None),
		(rur_density, 'Rural housing density',     'q_r·H_r/Y', None),
		# --- Production ---
		(I_Y,         'Investment/GDP',            'ratio', None),
		(quIHu_Y,     'Urban construction/GDP',    'ratio', None),
		(qrIHr_Y,     'Rural construction/GDP',    'ratio', None),
		# --- Prices ---
		(w,           'Real wage',                 'w',     None),
		(r,           'Real interest rate',        'r',     None),
		(q_u,         'Urban house price',         'q_u',   None),
		(q_r,         'Rural house price',         'q_r',   None),
	]

	n_panels = len(specs)
	ncols = 4
	nrows = int(np.ceil(n_panels / ncols))

	fig, axes = plt.subplots(nrows, ncols, figsize=(ncols * 4, nrows * 3))
	axes = axes.ravel()

	for ax, (series, title, ylabel, _) in zip(axes, specs):
		has_model = series is not None and np.any(np.isfinite(series))
		has_data  = data_overlay is not None and title in data_overlay
		if not has_model and not has_data:
			ax.set_title(title, fontsize=9)
			ax.text(0.5, 0.5, 'n/a', ha='center', va='center', transform=ax.transAxes,
			        color='grey', fontsize=10)
			ax.axis('off')
			continue
		if has_model:
			ax.plot(t_grid, series, lw=1.8, color='steelblue', label='model')
		if has_data:
			d_years, d_vals = data_overlay[title]
			ax.scatter(d_years, d_vals, s=18, color='firebrick', zorder=5,
			           label='data', linewidths=0)
		if has_model and has_data:
			ax.legend(fontsize=7)
		ax.set_title(title, fontsize=9)
		ax.set_ylabel(ylabel, fontsize=8)
		ax.set_xlabel('year', fontsize=8)
		ax.tick_params(labelsize=7)
		ax.grid(True, alpha=0.3)

	# hide unused axes
	for ax in axes[n_panels:]:
		ax.axis('off')

	plt.suptitle('Transition path moments', fontsize=12, y=1.01)
	plt.tight_layout()
	plt.show()

	if include_clearing:
		clear_specs = [
			(_get('clearing_A'),   'Asset market residual'),
			(_get('clearing_L'),   'Labor market residual'),
			(_get('clearing_H_u'), 'Urban housing residual'),
			(_get('clearing_H_r'), 'Rural housing residual'),
		]

		fig, axes = plt.subplots(2, 2, figsize=(12, 7))
		axes = axes.ravel()

		for ax, (s, title) in zip(axes, clear_specs):
			if s is None:
				ax.set_title(title); ax.axis('off'); continue
			ax.plot(t_grid, s, lw=1.8, color='firebrick')
			ax.axhline(0.0, color='k', ls='--', lw=1, alpha=0.6)
			ax.set_title(title, fontsize=10)
			ax.set_xlabel('year', fontsize=9)
			ax.tick_params(labelsize=8)
			ax.grid(True, alpha=0.3)

		plt.tight_layout()
		plt.show()


def plot_hh_jacobians(
	self,
	mode='aggregate',
	age=0,
	shock_dates=None,
	age_jac_results=None,
	percent_dev=False,
	dx=1e-4,
	age_horizon=None,
	selected_shock_ages=None,
	backward_window=6,
	do_print=False,
	t_max=None,
	common_color_scale=True,
	cmap='RdBu_r',
	outputs=None,
	inputs=None,
):
	"""Plot household Jacobians in an (outputs × inputs) grid.

	By default rows are auto-detected from what's actually populated on the
	model — ``self.jac_hh`` for mode='aggregate', or ``self.age_jac_hh``
	(or ``age_jac_results``) for mode='age-specific'. This avoids ValueError
	when some outputs aren't built (in this SOE pipeline ``compute_jacobians_
	complete`` builds H_u_hh and H_r_hh only; A_hh / C_hh are not stored).
	Columns default to ``self.inputs_hh``.

	Parameters:
	- mode: 'aggregate' or 'age-specific'
	- age: age index used when mode='age-specific'
	- shock_dates: list of shock columns s to plot
	- age_jac_results: optional dict keyed by (outputname, inputname) with precomputed
	  results from compute_age_specific_hh_jacobians
	- t_max: optional horizon for plotting (default: full Jacobian size)
	- outputs: optional list of output names; default = whatever is populated.
	  Pass e.g. ['H_u_hh','H_r_hh'] to force a specific subset.
	- inputs: optional list of input names; default = ``self.inputs_hh``.

	Important:
	- This function is read-only: it never recomputes Jacobians.
	- For mode='aggregate', it reads self.jac_hh.
	- For mode='age-specific', pass age_jac_results or store one on the model
	  as self.age_jac_hh.

	Returns:
	- J_dict: dict keyed by (outputname, inputname) with plotted Jacobian matrices
	- age_jac_results: dict of age-specific Jacobian results used
	"""

	mode = str(mode).strip().lower()
	if mode not in ('aggregate', 'age-specific'):
		raise ValueError("mode must be either 'aggregate' or 'age-specific'")

	if inputs is None:
		inputs = list(self.inputs_hh)
	else:
		inputs = list(inputs)

	# Auto-detect outputs from whatever is actually populated on the model.
	if outputs is None:
		if mode == 'aggregate':
			jh = getattr(self, 'jac_hh', None) or {}
			present = {o for (o, i) in jh.keys() if i in inputs}
		else:
			ajh = getattr(self, 'age_jac_hh', None) or (age_jac_results or {})
			present = {o for (o, i) in ajh.keys() if i in inputs}
		canonical = ['A_hh', 'C_hh', 'H_u_hh', 'H_r_hh']
		outputs = [o for o in canonical if o in present]
		if not outputs:
			raise ValueError(
				f"No household Jacobians found on the model for any of "
				f"{canonical} × {inputs}. Run "
				f"`transition.compute_jacobians_complete(model)` first."
			)
	else:
		outputs = list(outputs)

	# Backward-compatible but intentionally unused in plotting-only mode.
	_ = (dx, age_horizon, selected_shock_ages, backward_window, do_print, common_color_scale, cmap)

	def _scale_to_percent_dev(J, outputname, inputname):
		"""Scale dY/dX into % deviation of Y for 1% shock in X.

		If percent_dev=True, transforms:
		  dY/dX  ->  (ss_X / ss_Y) * dY/dX
		so a 1% change in X implies a (ss_X/ss_Y)*dY/dX percent change in Y.
		"""
		if not percent_dev:
			return J
		if not hasattr(self, 'ss'):
			return J
		ss = self.ss
		if (not hasattr(ss, outputname)) or (not hasattr(ss, inputname)):
			return J
		try:
			y_ss = float(getattr(ss, outputname))
			x_ss = float(getattr(ss, inputname))
		except Exception:
			return J
		if (not np.isfinite(y_ss)) or (not np.isfinite(x_ss)):
			return J
		if abs(y_ss) < 1e-14 or abs(x_ss) < 1e-14:
			return J
		return np.asarray(J, dtype=float) * (x_ss / y_ss)

	if age_jac_results is None:
		age_jac_results = getattr(self, 'age_jac_hh', None)
		if age_jac_results is None:
			age_jac_results = {}

	def _get_matrix(outputname, inputname):
		pair = (outputname, inputname)

		if mode == 'aggregate':
			if not hasattr(self, 'jac_hh') or pair not in self.jac_hh:
				raise ValueError(
					"Missing aggregate household Jacobians on model.jac_hh. "
					"Compute them first, then call plot_hh_jacobians(mode='aggregate')."
				)
			return np.asarray(self.jac_hh[pair])

		if pair not in age_jac_results:
			raise ValueError(
				"Missing age-specific Jacobians for pair "
				f"({outputname}, {inputname}). Provide age_jac_results, or rerun "
				"compute_jacobians_complete(...) to populate self.age_jac_hh."
			)

		res_pair = age_jac_results[pair]
		if 'J_by_age' not in res_pair:
			raise ValueError(
				"age_jac_results entries must contain key 'J_by_age'."
			)

		J_by_age = res_pair['J_by_age']
		A = len(J_by_age)
		age_idx = int(age)
		if age_idx < 0 or age_idx >= A:
			raise ValueError(f'age must be in [0, {A-1}]')
		return np.asarray(J_by_age[age_idx])

	J_dict = {}
	for outputname in outputs:
		for inputname in inputs:
			J_raw = _get_matrix(outputname, inputname)
			J = _scale_to_percent_dev(J_raw, outputname, inputname)
			if J.ndim != 2 or J.shape[0] != J.shape[1]:
				raise ValueError(f'Jacobian for ({outputname}, {inputname}) must be square 2D.')
			J_dict[(outputname, inputname)] = J

	T_plot = min(J.shape[0] for J in J_dict.values())
	if t_max is None:
		t_max = T_plot
	else:
		t_max = int(max(1, min(int(t_max), T_plot)))

	title_mode = 'Aggregate' if mode == 'aggregate' else f'Age-specific (age={int(age)})'

	n_out = len(outputs)
	n_in  = len(inputs)
	fig, axes = plt.subplots(n_out, n_in, figsize=(n_in * 4, n_out * 3.5), sharex=True)
	axes = np.array(axes).reshape(n_out, n_in)

	if shock_dates is None:
		cand = [0, t_max // 4, t_max // 2, (3 * t_max) // 4]
		shock_dates = sorted(set(int(s) for s in cand if 0 <= int(s) < t_max))
	else:
		shock_dates = sorted(set(int(s) for s in shock_dates if 0 <= int(s) < t_max))
	if len(shock_dates) == 0:
		raise ValueError('No valid shock_dates in plotting range.')

	t_grid = np.arange(t_max)
	ylab = 'Derivative'
	if percent_dev:
		ylab = '% dev. in output (from 1% input shock)'

	for i_out, outputname in enumerate(outputs):
		for j_in, inputname in enumerate(inputs):
			ax = axes[i_out, j_in]
			Jv = J_dict[(outputname, inputname)][:t_max, :t_max]
			for s in shock_dates:
				ax.plot(t_grid, Jv[:, s], lw=1.5, label=f's={s}')
			ax.axhline(0.0, color='k', ls='--', lw=0.8, alpha=0.7)
			ax.set_title(f'{outputname} wrt {inputname}', fontsize=9)
			ax.grid(True, alpha=0.25)

	for j_in in range(n_in):
		axes[n_out - 1, j_in].set_xlabel('Response date t')
	for i_out in range(n_out):
		axes[i_out, 0].set_ylabel(ylab)

	axes[0, 0].legend(fontsize=8, loc='best')

	fig.suptitle(f'{title_mode} Household Jacobians (all output-input pairs)', y=0.995)
	plt.tight_layout(rect=[0, 0, 1, 0.98])
	plt.show()

	return J_dict, age_jac_results







def plot_irf(irf, title='', t_max=None, pct=True, pct_floor=1e-8, ylim=None):
    """Plot linearised IRFs returned by linear_irf().

    2x2 figure:
        (0,0) Housing prices                            q_u, q_r           (GE)
        (0,1) Total housing demand                      H_u_hh, H_r_hh     (GE + PE)
        (1,0) Aggregate consumption + goods output      C_hh, Y            (GE)
        (1,1) HH savings + capital                      A_hh, K            (GE)

    PE lines are only drawn where the IRF dict contains a *_pe key. Every
    other panel shows the GE response alone.

    Args:
        irf      : dict from linear_irf(). May contain irf['_base'] with
                   baseline scalars for each plotted key — required when
                   pct=True so the y-axis can be expressed as % of SS.
        title    : optional suptitle.
        t_max    : truncate the x-axis at t_max.
        pct      : if True, plot 100 * (level-deviation) / baseline. Falls
                   back to raw level for any series whose baseline is missing
                   or smaller than `pct_floor` in absolute value.
        pct_floor: |baseline| below this is treated as "no meaningful base",
                   and that series is plotted in levels instead.
        ylim     : per-panel y-axis limits. Accepted forms:
                   - None  (default)      → matplotlib auto on every panel.
                   - (lo, hi)             → same limits on all four panels.
                   - list/tuple of length 4: per-panel limits in the layout
                     order [(0,0), (0,1), (1,0), (1,1)]. Each entry may be
                     None (auto for that panel) or a (lo, hi) tuple.
                   - dict keyed by panel title (e.g.
                     ``{'Housing prices': (-5, 5), 'Savings & capital': None}``)
                     or by linear panel index 0..3. Missing entries → auto.
    """
    import matplotlib.pyplot as plt
    import numpy as np

    fig, axes = plt.subplots(2, 2, figsize=(12, 7))
    panels = [
        (0, 0, 'Housing prices',          [('urban',       'q_u'),    ('rural',         'q_r')]),
        (0, 1, 'Total housing demand',    [('urban',       'H_u_hh'), ('rural',         'H_r_hh')]),
        (1, 0, 'Consumption & output',    [('consumption', 'C_hh'),   ('goods output',  'Y')]),
        (1, 1, 'Savings & capital',       [('HH savings',  'A_hh'),   ('capital stock', 'K')]),
    ]
    base_map = irf.get('_base', {}) if pct else {}

    def _scale(key, y):
        if not pct:
            return y, ''
        b = base_map.get(key, np.nan)
        if not np.isfinite(b) or abs(b) < pct_floor:
            return y, ''                              # silently fall back to level
        return 100.0 * y / b, ' (% SS)'

    def _resolve_ylim(panel_idx, panel_title):
        """Pick the (lo, hi) for one panel from the user's `ylim` argument.

        Returns None to mean "use matplotlib auto for this panel".
        """
        if ylim is None:
            return None
        if isinstance(ylim, dict):
            return ylim.get(panel_title, ylim.get(panel_idx))
        if isinstance(ylim, (list, tuple)):
            # Disambiguate a single (lo, hi) from a list-of-4 panel limits.
            is_single = (
                len(ylim) == 2
                and not isinstance(ylim[0], (list, tuple, dict))
                and not isinstance(ylim[1], (list, tuple, dict))
            )
            if is_single:
                return tuple(ylim)
            if panel_idx < len(ylim):
                return ylim[panel_idx]
        return None

		
    def _series_color(label, i):
        """urban -> blue, rural -> red; other (non-regional) series -> fallback."""
        lab = label.lower()
        if 'urban' in lab:
            return '#003C8F'   # blue
        if 'rural' in lab:
            return '#C92828'   # red
        # non-regional panels: 1st series blue, 2nd red (same palette)
        return ('#003C8F', '#C92828')[i % 2]

    ylabel_used = False
    for panel_idx, (r, c, ptitle, series) in enumerate(panels):
        ax = axes[r, c]
        suffix = ''
        for i, (label, key) in enumerate(series):
            color = _series_color(label, i)
            if key in irf:
                y, suffix = _scale(key, np.asarray(irf[key]))
                ax.plot(y, label=f'{label} (GE)', color=color, lw=1.8)
            if (key + '_pe') in irf:
                y_pe, _ = _scale(key, np.asarray(irf[key + '_pe']))
                ax.plot(y_pe, label=f'{label} (PE)',
                        color=color, lw=1.2, linestyle='--', alpha=0.85)
        ax.set_title(ptitle + suffix)
        ax.set_xlabel('t')
        if pct and suffix and not ylabel_used:
            ax.set_ylabel('% deviation from SS')
            ylabel_used = True
        ax.axhline(0, color='k', lw=0.5)
        ax.legend(fontsize=8, frameon=False)
        if t_max is not None:
            ax.set_xlim(0, t_max)
        yl = _resolve_ylim(panel_idx, ptitle)
        if yl is not None:
            ax.set_ylim(*yl)
    if title:
        fig.suptitle(title, fontsize=13, y=1.00)
    plt.tight_layout()
    plt.show()
    return fig




def plot_comparative_statics_tenure_supply_demand(
	model_baseline=None,
	shock='kappa',
	shock_size=0.05,
	n_points=15,
	rel_price_span=(2.0, 0.5),
	data=None,
	title=None,
	do_print=False,
):
	"""Plot the 2x2 tenure supply/demand diagrams (baseline vs shock).

	The heavy HH-solve computation lives in
	`steady_state.comparative_statics_tenure_supply_demand`.  This function
	only does the plotting.  Two call patterns:

	1. **Compute-and-plot** — pass `model_baseline` (and optionally the
	   `shock` / `shock_size` kwargs).  The compute function is run, the data
	   dict is built, then plotted.

	2. **Plot-only (recommended for iterating on plot styling)** — first
	   compute the data dict once:
	       data = steady_state.comparative_statics_tenure_supply_demand(
	                  model_baseline, shock='kappa', shock_size=0.05)
	   then call
	       plots.plot_comparative_statics_tenure_supply_demand(data=data,
	                                                            title='...')
	   to re-plot without re-solving anything.

	Layout (2x2):
	- Left column: urban  /  Right column: rural
	- Top row: ownership relative prices  (y = q_u/q_r or q_r/q_u)
	- Bottom row: rent relative prices    (y = f_u/f_r or f_r/f_u)
	In each panel: demand owned (red), demand rental (green),
	demand total (dark blue), supply (black).  Baseline = solid,
	shocked = dashed.

	Parameters
	----------
	data : dict, optional
	    Precomputed data dict returned by
	    `steady_state.comparative_statics_tenure_supply_demand`.  When given,
	    `model_baseline` and the shock-related kwargs are ignored.
	model_baseline, shock, shock_size, rel_price_span, n_points, do_print :
	    Forwarded to the compute function when `data is None`.
	title : str, optional
	    Suptitle.  Defaults to a one-line auto-generated description.
	"""
	# Resolve `data`: either use the precomputed dict, or compute it now.
	if data is None:
		if model_baseline is None:
			raise ValueError(
				'Either `data=` (precomputed dict from '
				'steady_state.comparative_statics_tenure_supply_demand) '
				'or `model_baseline=` must be provided.'
			)
		import steady_state as ss_mod
		data = ss_mod.comparative_statics_tenure_supply_demand(
			model_baseline,
			shock=shock,
			shock_size=shock_size,
			n_points=n_points,
			rel_price_span=rel_price_span,
			do_print=do_print,
		)

	# Light input sanitisation for the plot-only knobs.
	if not isinstance(title, str):
		title = None
	# --- Unpack the data dict ------------------------------------------------
	# shock_desc is optional — older / pickled data dicts may not have it,
	# in which case the suptitle just falls back to a generic label below.
	shock_desc       = data.get('shock_desc', data.get('shock', 'shock'))
	# Per-panel price-ratio grids (new schema).  Fall back to legacy single
	# grids if loading a pre-refactor data dict — rural panel will look wrong
	# in that case (vertical supply) but won't crash.
	ratio_q_grid_u   = np.asarray(data.get('ratio_q_grid_u', data['ratio_q_grid']))
	ratio_q_grid_r   = np.asarray(data.get('ratio_q_grid_r', 1.0 / ratio_q_grid_u))
	ratio_f_grid_u   = np.asarray(data.get('ratio_f_grid_u', data['ratio_f_grid']))
	ratio_f_grid_r   = np.asarray(data.get('ratio_f_grid_r', 1.0 / ratio_f_grid_u))

	q_d_u_owner_base  = data['q_d_u_owner_base'];  q_d_u_owner_shock  = data['q_d_u_owner_shock']
	q_d_r_owner_base  = data['q_d_r_owner_base'];  q_d_r_owner_shock  = data['q_d_r_owner_shock']
	q_d_u_rent_base   = data['q_d_u_rent_base'];   q_d_u_rent_shock   = data['q_d_u_rent_shock']
	q_d_r_rent_base   = data['q_d_r_rent_base'];   q_d_r_rent_shock   = data['q_d_r_rent_shock']
	q_d_u_total_base  = data['q_d_u_total_base'];  q_d_u_total_shock  = data['q_d_u_total_shock']
	q_d_r_total_base  = data['q_d_r_total_base'];  q_d_r_total_shock  = data['q_d_r_total_shock']

	f_d_u_owner_base  = data['f_d_u_owner_base'];  f_d_u_owner_shock  = data['f_d_u_owner_shock']
	f_d_r_owner_base  = data['f_d_r_owner_base'];  f_d_r_owner_shock  = data['f_d_r_owner_shock']
	f_d_u_rent_base   = data['f_d_u_rent_base'];   f_d_u_rent_shock   = data['f_d_u_rent_shock']
	f_d_r_rent_base   = data['f_d_r_rent_base'];   f_d_r_rent_shock   = data['f_d_r_rent_shock']
	f_d_u_total_base  = data['f_d_u_total_base'];  f_d_u_total_shock  = data['f_d_u_total_shock']
	f_d_r_total_base  = data['f_d_r_total_base'];  f_d_r_total_shock  = data['f_d_r_total_shock']

	q_s_u_base = data['q_s_u_base'];  q_s_u_shock = data['q_s_u_shock']
	q_s_r_base = data['q_s_r_base'];  q_s_r_shock = data['q_s_r_shock']
	f_s_u_base = data['f_s_u_base'];  f_s_u_shock = data['f_s_u_shock']
	f_s_r_base = data['f_s_r_base'];  f_s_r_shock = data['f_s_r_shock']

	# Drop into the unchanged plotting code below.
	# ------------------------------------------------------------------
	# 3. Plot curves (x=quantity, y=relative price)
	# ------------------------------------------------------------------
	if title is None:
		title = (
			f'Comparative statics (baseline vs shock): {shock_desc}'
		)

	fig, axes = plt.subplots(
		2,
		2,
		figsize=(9, 8),
		sharey=False,
		gridspec_kw={'wspace': 0.01, 'hspace': 0.20},
	)

	# Shared style across panels
	col_owned = '#C92828'   # red
	col_rent = '#1F7A1F'    # green
	col_supply = '#111111'  # black
	col_total = '#003C8F'   # dark blue

	# Top-left: Urban ownership-price diagram (q_u/q_r)
	ax = axes[0, 0]
	ax.plot(q_d_u_owner_base, ratio_q_grid_u, color=col_owned, lw=2.0, ls='-', label='Demand owned (baseline)')
	ax.plot(q_d_u_rent_base,  ratio_q_grid_u, color=col_rent, lw=2.0, ls='-', label='Demand rental (baseline)')
	ax.plot(q_d_u_total_base, ratio_q_grid_u, color=col_total, lw=2.0, ls='-', label='Demand total (baseline)')
	ax.plot(q_s_u_base, ratio_q_grid_u, color=col_supply, lw=1.0, ls='-', label='Supply (baseline)')
	ax.plot(q_d_u_owner_shock, ratio_q_grid_u, color=col_owned, lw=2.0, ls='--', label='Demand owned (shock)')
	ax.plot(q_d_u_rent_shock,  ratio_q_grid_u, color=col_rent, lw=2.0, ls='--', label='Demand rental (shock)')
	ax.plot(q_d_u_total_shock, ratio_q_grid_u, color=col_total, lw=2.0, ls='--', label='Demand total (shock)')
	ax.plot(q_s_u_shock, ratio_q_grid_u, color=col_supply, lw=1.0, ls='--', label='Supply (shock)')
	ax.set_title('Urban', fontsize=11)
	ax.set_ylabel('$q^u/q^r$')
	ax.grid(False)
	ax.set_box_aspect(1.0)

	# Top-right: Rural ownership-price diagram (q_r/q_u)
	ax = axes[0, 1]
	ax.plot(q_d_r_owner_base, ratio_q_grid_r, color=col_owned, lw=2.0, ls='-', label='Demand owned (baseline)')
	ax.plot(q_d_r_rent_base,  ratio_q_grid_r, color=col_rent, lw=2.0, ls='-', label='Demand rental (baseline)')
	ax.plot(q_d_r_total_base, ratio_q_grid_r, color=col_total, lw=2.2, ls='-', label='Demand total (baseline)')
	ax.plot(q_s_r_base, ratio_q_grid_r, color=col_supply, lw=1.0, ls='-', label='Supply (baseline)')
	ax.plot(q_d_r_owner_shock, ratio_q_grid_r, color=col_owned, lw=2.0, ls='--', label='Demand owned (shock)')
	ax.plot(q_d_r_rent_shock,  ratio_q_grid_r, color=col_rent, lw=2.0, ls='--', label='Demand rental (shock)')
	ax.plot(q_d_r_total_shock, ratio_q_grid_r, color=col_total, lw=2.2, ls='--', label='Demand total (shock)')
	ax.plot(q_s_r_shock, ratio_q_grid_r, color=col_supply, lw=1.0, ls='--', label='Supply (shock)')
	ax.set_title('Rural', fontsize=11)
	ax.set_ylabel('$q^r/q^u$')
	ax.grid(False)
	ax.set_box_aspect(1.0)

	# Bottom-left: Urban rent diagram (f_u/f_r)
	ax = axes[1, 0]
	ax.plot(f_d_u_rent_base, ratio_f_grid_u, color=col_rent, lw=2.0, ls='-', label='Demand rental (baseline)')
	ax.plot(f_d_u_owner_base, ratio_f_grid_u, color=col_owned, lw=2.0, ls='-', label='Demand owned (baseline)')
	ax.plot(f_d_u_total_base, ratio_f_grid_u, color=col_total, lw=2.2, ls='-', label='Demand total (baseline)')
	ax.plot(f_s_u_base, ratio_f_grid_u, color=col_supply, lw=1.0, ls='-', label='Supply (baseline)')
	ax.plot(f_d_u_rent_shock, ratio_f_grid_u, color=col_rent, lw=2.0, ls='--', label='Demand rental (shock)')
	ax.plot(f_d_u_owner_shock, ratio_f_grid_u, color=col_owned, lw=2.0, ls='--', label='Demand owned (shock)')
	ax.plot(f_d_u_total_shock, ratio_f_grid_u, color=col_total, lw=2.2, ls='--', label='Demand total (shock)')
	ax.plot(f_s_u_shock, ratio_f_grid_u, color=col_supply, lw=1.0, ls='--', label='Supply (shock)')
	ax.set_ylabel('$f^u/f^r$')
	ax.set_xlabel('Urban housing quantity')
	ax.grid(False)
	ax.set_box_aspect(1.0)

	# Bottom-right: Rural rent diagram (f_r/f_u)
	ax = axes[1, 1]
	ax.plot(f_d_r_rent_base, ratio_f_grid_r, color=col_rent, lw=2.0, ls='-', label='Demand rental (baseline)')
	ax.plot(f_d_r_owner_base, ratio_f_grid_r, color=col_owned, lw=2.0, ls='-', label='Demand owned (baseline)')
	ax.plot(f_d_r_total_base, ratio_f_grid_r, color=col_total, lw=2.2, ls='-', label='Demand total (baseline)')
	ax.plot(f_s_r_base, ratio_f_grid_r, color=col_supply, lw=1.0, ls='-', label='Supply (baseline)')
	ax.plot(f_d_r_rent_shock, ratio_f_grid_r, color=col_rent, lw=2.0, ls='--', label='Demand rental (shock)')
	ax.plot(f_d_r_owner_shock, ratio_f_grid_r, color=col_owned, lw=2.0, ls='--', label='Demand owned (shock)')
	ax.plot(f_d_r_total_shock, ratio_f_grid_r, color=col_total, lw=2.2, ls='--', label='Demand total (shock)')
	ax.plot(f_s_r_shock, ratio_f_grid_r, color=col_supply, lw=1.0, ls='--', label='Supply (shock)')
	ax.set_ylabel('$f^r/f^u$')
	ax.set_xlabel('Rural housing quantity')
	ax.grid(False)
	ax.set_box_aspect(1.0)

	# Per-panel axis limits.  The y-axis is taken from THIS panel's own
	# price-ratio grid, so the baseline ratio sits in the middle of the panel
	# (not pinned to the bottom of a shared [0.5,5] window) and the demand
	# schedules fill the panel.  The x-axis runs from 0 to just past the
	# largest demand quantity.
	#
	# Supply (∝ q^pow, pow=3) is deliberately NOT used to set x_max: it would
	# blow the horizontal scale up by an order of magnitude and crush all the
	# demand detail.  Because the SS solver clears each market on exactly this
	# supply schedule (find_ss_prices: q = q_zp(H_hh) ⇔ supply(q)=demand(q)),
	# the supply curve still crosses total demand at the steady-state point
	# inside this window; above the baseline ratio it simply exits the right
	# edge.
	def _set_axes(ax, demand_x_vals, y_grid, x_headroom=1.15, y_pad_frac=0.02):
		dx = np.concatenate([np.asarray(v).ravel() for v in demand_x_vals])
		dx = dx[np.isfinite(dx)]
		y = np.asarray(y_grid).ravel()
		y = y[np.isfinite(y)]
		if dx.size > 0:
			xmax = x_headroom * float(np.max(dx))
			if not np.isfinite(xmax) or xmax <= 0.0:
				xmax = 10.0
			ax.set_xlim(0.0, xmax)
		if y.size > 0:
			ymin = float(np.min(y))
			ymax = float(np.max(y))
			ypad = y_pad_frac * max(ymax - ymin, 1e-12)
			ax.set_ylim(ymin - ypad, ymax + ypad)
		ax.margins(x=0.0, y=0.0)

	_set_axes(
		axes[0, 0],
		[q_d_u_owner_base, q_d_u_rent_base, q_d_u_total_base,
		 q_d_u_owner_shock, q_d_u_rent_shock, q_d_u_total_shock],
		ratio_q_grid_u,
	)
	_set_axes(
		axes[0, 1],
		[q_d_r_owner_base, q_d_r_rent_base, q_d_r_total_base,
		 q_d_r_owner_shock, q_d_r_rent_shock, q_d_r_total_shock],
		ratio_q_grid_r,
	)
	_set_axes(
		axes[1, 0],
		[f_d_u_owner_base, f_d_u_rent_base, f_d_u_total_base,
		 f_d_u_owner_shock, f_d_u_rent_shock, f_d_u_total_shock],
		ratio_f_grid_u,
	)
	_set_axes(
		axes[1, 1],
		[f_d_r_owner_base, f_d_r_rent_base, f_d_r_total_base,
		 f_d_r_owner_shock, f_d_r_rent_shock, f_d_r_total_shock],
		ratio_f_grid_r,
	)

	handles, labels = axes[0, 0].get_legend_handles_labels()
	fig.legend(
		handles,
		labels,
		loc='upper center',
		ncol=4,
		frameon=False,
		bbox_to_anchor=(0.5, 0.975),
		fontsize=10,
	)
	fig.suptitle(title, fontsize=14, y=0.995)
	fig.tight_layout(rect=[0.0, 0.0, 1.0, 0.84], w_pad=0.01, h_pad=0.15)
	plt.show()
	plt.close(fig)


def load_datagraphs(filepath, start_year=1992):
	"""Load empirical data from datagraphs.xlsx.

	Returns a dict mapping panel title (as used in plot_transition_paths) to
	a tuple (years, values) of numpy arrays.  Rows with all-NaN values are
	skipped.  The year range is inferred from the number of data rows.

	Parameters
	----------
	filepath   : str – path to the xlsx file
	start_year : int – calendar year of the first data row (default 1992)
	"""
	import openpyxl

	def _norm_header(x):
		# Robust to whitespace/case differences in Excel headers.
		return ' '.join(str(x).strip().split()).lower()

	# Mapping: xlsx column header → plot panel title
	_COL_TO_TITLE = {
		'Capital/GDP':            'Capital/GDP',
		'Urban housing/GDP':      'Urban housing/GDP',
		'Rural housing/GDP':      'Rural housing/GDP',
		'Urban housing wealth':   'Urban housing wealth',
		'Rural housing wealth':   'Rural housing wealth',
		'Top 1':                  'Top 1% wealth share',
		'Top 10':                 'Top 10% wealth share',
		'Middle 40':              'Middle 40% wealth share',
		'Bottom 50':              'Bottom 50% wealth share',
		'Urban price/rural price':'Urban/rural price ratio',
		'Urban population share': 'Urban population share',
		'Urban housing density':  'Urban housing density',
		'Rural housing density':  'Rural housing density',
		'Investment/GDP':         'Investment/GDP',
		'Urban construction/GDP': 'Urban construction/GDP',
		'Rural construction/GDP': 'Rural construction/GDP',
		'Real rate':              'Real rate',
		'Net Foreign Assets':     'Net Foreign Assets/GDP',
		'Net Foreign Assets/GDP': 'Net Foreign Assets/GDP',
		'Share of renters - Urban Region': 'Share of renters - Urban region',
		'Share of renters - Urban region': 'Share of renters - Urban region',
		'Share of renters - Rural Region': 'Share of renters - Rural region',
		'Share of renters - Rural region': 'Share of renters - Rural region',
	}
	_COL_TO_TITLE_NORM = {_norm_header(k): v for k, v in _COL_TO_TITLE.items()}

	wb  = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
	ws  = wb.active
	rows = list(ws.iter_rows(values_only=True))
	wb.close()

	header   = [str(c).strip() if c is not None else '' for c in rows[0]]
	data_rows = rows[1:]
	n_rows    = len(data_rows)
	years     = np.arange(start_year, start_year + n_rows)

	result = {}
	for col_idx, col_name in enumerate(header):
		title = _COL_TO_TITLE_NORM.get(_norm_header(col_name))
		if title is None:
			continue
		vals = np.array(
			[float(r[col_idx]) if (r[col_idx] is not None and
			                        str(r[col_idx]).strip() != '') else np.nan
			 for r in data_rows],
			dtype=float,
		)
		if np.any(np.isfinite(vals)):
			result[title] = (years, vals)

	return result


def plot_transition_paths(self, T_plot=None, include_clearing=True, start_year=1992,
                          data_overlay=None):
	"""Plot transition paths for the moments listed in Table X of the paper.

	Groups:
	  Wealth ratios      : Capital/GDP, Urban housing/GDP, Rural housing/GDP
	  Wealth inequality  : Top 1%, Top 10%, Bottom 50% (asset share)
	  Housing            : Urban/rural price ratio, Urban pop share,
	                       Urban housing density, Rural housing density
	  Production         : Investment/GDP, Urban construction/GDP, Rural construction/GDP
	  Prices             : Real wage, Real interest rate, Urban house price, Rural house price

	Parameters
	----------
	T_plot        : int, optional – number of periods to plot (default: full horizon)
	include_clearing : bool – if True, append a market-clearing residual panel
	start_year    : int – calendar year of t=0 for x-axis labels (default 1992)
	data_overlay  : dict or str, optional – if a str, treated as a path to an xlsx
	                file loaded via load_datagraphs(); if a dict (from load_datagraphs),
	                used directly. Maps panel title → (years, values) and overlays
	                the empirical series as dots on each matching panel.
	"""

	if not hasattr(self, 'path'):
		print("Path object not available on model.")
		return

	# Resolve data_overlay
	if isinstance(data_overlay, str):
		data_overlay = load_datagraphs(data_overlay, start_year=start_year)

	par = self.par
	ss  = self.ss
	path = self.path

	T = par.T if hasattr(par, 'T') else 1
	if T_plot is None:
		T_plot = T
	T_plot = int(max(1, min(T_plot, T)))
	t_grid = np.arange(T_plot) + start_year

	def _get(name):
		if not hasattr(path, name):
			return None
		arr = np.asarray(getattr(path, name))
		return arr[:T_plot, 0] if arr.ndim == 2 else arr[:T_plot]

	def _ss(name):
		return float(getattr(ss, name)) if hasattr(ss, name) else np.nan

	# ------------------------------------------------------------------
	# Build derived series
	# ------------------------------------------------------------------
	Y    = _get('Y')
	K    = _get('K')
	q_u  = _get('q_u')
	q_r  = _get('q_r')
	H_u  = _get('H_u')
	H_r  = _get('H_r')
	IH_u = _get('IH_u')
	IH_r = _get('IH_r')
	I    = _get('I')
	w    = _get('w')
	r    = _get('r')
	A_hh = _get('A_hh')

	#GDP definition as per equation (3.21):
	Y_tot = Y + q_u * IH_u + q_r * IH_r

	# Safe denominator
	Y_safe = np.where((Y_tot is not None) & np.isfinite(Y_tot) & (Y_tot > 0), Y_tot, np.nan) if Y_tot is not None else None

	def _ratio(num, den_safe, label):
		if num is None or den_safe is None:
			return None
		return np.where(np.isfinite(num) & np.isfinite(den_safe), num / den_safe, np.nan)

	# Wealth ratios
	K_Y    = _ratio(K,               Y_safe, 'K/Y')
	quHu_Y = _ratio(q_u * H_u if (q_u is not None and H_u is not None) else None, Y_safe, 'quHu/Y')
	qrHr_Y = _ratio(q_r * H_r if (q_r is not None and H_r is not None) else None, Y_safe, 'qrHr/Y')

	# Housing ratios
	price_ratio = (q_u / np.where(q_r > 0, q_r, np.nan)
	               if (q_u is not None and q_r is not None) else None)

	# Urban population share: convert household housing demand back to mass
	# of households via h_u, h_r so it matches the calibration moment
	# (D[:,1,:].sum() + D[:,3,:].sum()) / total_mass.
	H_u_hh = _get('H_u_hh')
	H_r_hh = _get('H_r_hh')
	if H_u_hh is not None and H_r_hh is not None:
		N_u = H_u_hh / par.h_u
		N_r = H_r_hh / par.h_r
		N_total = N_u + N_r
		urb_share = np.where(N_total > 0, N_u / N_total, np.nan)
	else:
		urb_share = None

	# Housing density: price * stock / GDP (value of housing per unit GDP)
	urb_density = _ratio(H_u if (q_u is not None and H_u is not None) else None,
	                     par.X_u, 'urb_density')
	rur_density = _ratio(H_r if (q_r is not None and H_r is not None) else None,
	                     par.X_r, 'rur_density')

	# Investment and construction ratios
	I_Y     = _ratio(I,    Y_safe, 'I/Y')
	quIHu_Y = _ratio(q_u * IH_u if (q_u is not None and IH_u is not None) else None, Y_safe, 'quIHu/Y')
	qrIHr_Y = _ratio(q_r * IH_r if (q_r is not None and IH_r is not None) else None, Y_safe, 'qrIHr/Y')

	# Wealth inequality: asset-share quantiles from simulated path.D
	# Populated by transition.compute_wealth_inequality(); skip gracefully if absent.
	top1_share  = _get('top1_share')
	top10_share = _get('top10_share')
	bot50_share = _get('bot50_share')

	# ------------------------------------------------------------------
	# Panel layout: 4 groups, each drawn in a separate figure row
	# ------------------------------------------------------------------
	specs = [
		# (series, title, y-label, ss_val_name)
		# --- Wealth ratios ---
		(K_Y,        'Capital/GDP',               'ratio', None),
		(quHu_Y,     'Urban housing/GDP',          'ratio', None),
		(qrHr_Y,     'Rural housing/GDP',          'ratio', None),
		# --- Wealth inequality ---
		(top1_share,  'Top 1% wealth share',       'share', None),
		(top10_share, 'Top 10% wealth share',      'share', None),
		(bot50_share, 'Bottom 50% wealth share',   'share', None),
		# --- Housing ---
		(price_ratio, 'Urban/rural price ratio',   'q_u/q_r', None),
		(urb_share,   'Urban population share',    'share',   None),
		(urb_density, 'Urban housing density',     'q_u·H_u/Y', None),
		(rur_density, 'Rural housing density',     'q_r·H_r/Y', None),
		# --- Production ---
		(I_Y,         'Investment/GDP',            'ratio', None),
		(quIHu_Y,     'Urban construction/GDP',    'ratio', None),
		(qrIHr_Y,     'Rural construction/GDP',    'ratio', None),
		# --- Prices ---
		(w,           'Real wage',                 'w',     None),
		(r,           'Real interest rate',        'r',     None),
		(q_u,         'Urban house price',         'q_u',   None),
		(q_r,         'Rural house price',         'q_r',   None),
	]

	n_panels = len(specs)
	ncols = 4
	nrows = int(np.ceil(n_panels / ncols))

	fig, axes = plt.subplots(nrows, ncols, figsize=(ncols * 4, nrows * 3))
	axes = axes.ravel()

	for ax, (series, title, ylabel, _) in zip(axes, specs):
		has_model = series is not None and np.any(np.isfinite(series))
		has_data  = data_overlay is not None and title in data_overlay
		if not has_model and not has_data:
			ax.set_title(title, fontsize=9)
			ax.text(0.5, 0.5, 'n/a', ha='center', va='center', transform=ax.transAxes,
			        color='grey', fontsize=10)
			ax.axis('off')
			continue
		if has_model:
			ax.plot(t_grid, series, lw=1.8, color='steelblue', label='model')
		if has_data:
			d_years, d_vals = data_overlay[title]
			ax.scatter(d_years, d_vals, s=18, color='firebrick', zorder=5,
			           label='data', linewidths=0)
		if has_model and has_data:
			ax.legend(fontsize=7)
		ax.set_title(title, fontsize=9)
		ax.set_ylabel(ylabel, fontsize=8)
		ax.set_xlabel('year', fontsize=8)
		ax.tick_params(labelsize=7)
		ax.grid(True, alpha=0.3)

	# hide unused axes
	for ax in axes[n_panels:]:
		ax.axis('off')

	plt.suptitle('Transition path moments', fontsize=12, y=1.01)
	plt.tight_layout()
	plt.show()

	if include_clearing:
		clear_specs = [
			(_get('clearing_A'),   'Asset market residual'),
			(_get('clearing_L'),   'Labor market residual'),
			(_get('clearing_H_u'), 'Urban housing residual'),
			(_get('clearing_H_r'), 'Rural housing residual'),
		]

		fig, axes = plt.subplots(2, 2, figsize=(12, 7))
		axes = axes.ravel()

		for ax, (s, title) in zip(axes, clear_specs):
			if s is None:
				ax.set_title(title); ax.axis('off'); continue
			ax.plot(t_grid, s, lw=1.8, color='firebrick')
			ax.axhline(0.0, color='k', ls='--', lw=1, alpha=0.6)
			ax.set_title(title, fontsize=10)
			ax.set_xlabel('year', fontsize=9)
			ax.tick_params(labelsize=8)
			ax.grid(True, alpha=0.3)

		plt.tight_layout()
		plt.show()


def plot_hh_jacobians(
	self,
	mode='aggregate',
	age=0,
	shock_dates=None,
	age_jac_results=None,
	percent_dev=False,
	dx=1e-4,
	age_horizon=None,
	selected_shock_ages=None,
	backward_window=6,
	do_print=False,
	t_max=None,
	common_color_scale=True,
	cmap='RdBu_r',
	outputs=None,
	inputs=None,
):
	"""Plot household Jacobians in an (outputs × inputs) grid.

	By default rows are auto-detected from what's actually populated on the
	model — ``self.jac_hh`` for mode='aggregate', or ``self.age_jac_hh``
	(or ``age_jac_results``) for mode='age-specific'. This avoids ValueError
	when some outputs aren't built (in this SOE pipeline ``compute_jacobians_
	complete`` builds H_u_hh and H_r_hh only; A_hh / C_hh are not stored).
	Columns default to ``self.inputs_hh``.

	Parameters:
	- mode: 'aggregate' or 'age-specific'
	- age: age index used when mode='age-specific'
	- shock_dates: list of shock columns s to plot
	- age_jac_results: optional dict keyed by (outputname, inputname) with precomputed
	  results from compute_age_specific_hh_jacobians
	- t_max: optional horizon for plotting (default: full Jacobian size)
	- outputs: optional list of output names; default = whatever is populated.
	  Pass e.g. ['H_u_hh','H_r_hh'] to force a specific subset.
	- inputs: optional list of input names; default = ``self.inputs_hh``.

	Important:
	- This function is read-only: it never recomputes Jacobians.
	- For mode='aggregate', it reads self.jac_hh.
	- For mode='age-specific', pass age_jac_results or store one on the model
	  as self.age_jac_hh.

	Returns:
	- J_dict: dict keyed by (outputname, inputname) with plotted Jacobian matrices
	- age_jac_results: dict of age-specific Jacobian results used
	"""

	mode = str(mode).strip().lower()
	if mode not in ('aggregate', 'age-specific'):
		raise ValueError("mode must be either 'aggregate' or 'age-specific'")

	if inputs is None:
		inputs = list(self.inputs_hh)
	else:
		inputs = list(inputs)

	# Auto-detect outputs from whatever is actually populated on the model.
	if outputs is None:
		if mode == 'aggregate':
			jh = getattr(self, 'jac_hh', None) or {}
			present = {o for (o, i) in jh.keys() if i in inputs}
		else:
			ajh = getattr(self, 'age_jac_hh', None) or (age_jac_results or {})
			present = {o for (o, i) in ajh.keys() if i in inputs}
		canonical = ['A_hh', 'C_hh', 'H_u_hh', 'H_r_hh']
		outputs = [o for o in canonical if o in present]
		if not outputs:
			raise ValueError(
				f"No household Jacobians found on the model for any of "
				f"{canonical} × {inputs}. Run "
				f"`transition.compute_jacobians_complete(model)` first."
			)
	else:
		outputs = list(outputs)

	# Backward-compatible but intentionally unused in plotting-only mode.
	_ = (dx, age_horizon, selected_shock_ages, backward_window, do_print, common_color_scale, cmap)

	def _scale_to_percent_dev(J, outputname, inputname):
		"""Scale dY/dX into % deviation of Y for 1% shock in X.

		If percent_dev=True, transforms:
		  dY/dX  ->  (ss_X / ss_Y) * dY/dX
		so a 1% change in X implies a (ss_X/ss_Y)*dY/dX percent change in Y.
		"""
		if not percent_dev:
			return J
		if not hasattr(self, 'ss'):
			return J
		ss = self.ss
		if (not hasattr(ss, outputname)) or (not hasattr(ss, inputname)):
			return J
		try:
			y_ss = float(getattr(ss, outputname))
			x_ss = float(getattr(ss, inputname))
		except Exception:
			return J
		if (not np.isfinite(y_ss)) or (not np.isfinite(x_ss)):
			return J
		if abs(y_ss) < 1e-14 or abs(x_ss) < 1e-14:
			return J
		return np.asarray(J, dtype=float) * (x_ss / y_ss)

	if age_jac_results is None:
		age_jac_results = getattr(self, 'age_jac_hh', None)
		if age_jac_results is None:
			age_jac_results = {}

	def _get_matrix(outputname, inputname):
		pair = (outputname, inputname)

		if mode == 'aggregate':
			if not hasattr(self, 'jac_hh') or pair not in self.jac_hh:
				raise ValueError(
					"Missing aggregate household Jacobians on model.jac_hh. "
					"Compute them first, then call plot_hh_jacobians(mode='aggregate')."
				)
			return np.asarray(self.jac_hh[pair])

		if pair not in age_jac_results:
			raise ValueError(
				"Missing age-specific Jacobians for pair "
				f"({outputname}, {inputname}). Provide age_jac_results, or rerun "
				"compute_jacobians_complete(...) to populate self.age_jac_hh."
			)

		res_pair = age_jac_results[pair]
		if 'J_by_age' not in res_pair:
			raise ValueError(
				"age_jac_results entries must contain key 'J_by_age'."
			)

		J_by_age = res_pair['J_by_age']
		A = len(J_by_age)
		age_idx = int(age)
		if age_idx < 0 or age_idx >= A:
			raise ValueError(f'age must be in [0, {A-1}]')
		return np.asarray(J_by_age[age_idx])

	J_dict = {}
	for outputname in outputs:
		for inputname in inputs:
			J_raw = _get_matrix(outputname, inputname)
			J = _scale_to_percent_dev(J_raw, outputname, inputname)
			if J.ndim != 2 or J.shape[0] != J.shape[1]:
				raise ValueError(f'Jacobian for ({outputname}, {inputname}) must be square 2D.')
			J_dict[(outputname, inputname)] = J

	T_plot = min(J.shape[0] for J in J_dict.values())
	if t_max is None:
		t_max = T_plot
	else:
		t_max = int(max(1, min(int(t_max), T_plot)))

	title_mode = 'Aggregate' if mode == 'aggregate' else f'Age-specific (age={int(age)})'

	n_out = len(outputs)
	n_in  = len(inputs)
	fig, axes = plt.subplots(n_out, n_in, figsize=(n_in * 4, n_out * 3.5), sharex=True)
	axes = np.array(axes).reshape(n_out, n_in)

	if shock_dates is None:
		cand = [0, t_max // 4, t_max // 2, (3 * t_max) // 4]
		shock_dates = sorted(set(int(s) for s in cand if 0 <= int(s) < t_max))
	else:
		shock_dates = sorted(set(int(s) for s in shock_dates if 0 <= int(s) < t_max))
	if len(shock_dates) == 0:
		raise ValueError('No valid shock_dates in plotting range.')

	t_grid = np.arange(t_max)
	ylab = 'Derivative'
	if percent_dev:
		ylab = '% dev. in output (from 1% input shock)'

	for i_out, outputname in enumerate(outputs):
		for j_in, inputname in enumerate(inputs):
			ax = axes[i_out, j_in]
			Jv = J_dict[(outputname, inputname)][:t_max, :t_max]
			for s in shock_dates:
				ax.plot(t_grid, Jv[:, s], lw=1.5, label=f's={s}')
			ax.axhline(0.0, color='k', ls='--', lw=0.8, alpha=0.7)
			ax.set_title(f'{outputname} wrt {inputname}', fontsize=9)
			ax.grid(True, alpha=0.25)

	for j_in in range(n_in):
		axes[n_out - 1, j_in].set_xlabel('Response date t')
	for i_out in range(n_out):
		axes[i_out, 0].set_ylabel(ylab)

	axes[0, 0].legend(fontsize=8, loc='best')

	fig.suptitle(f'{title_mode} Household Jacobians (all output-input pairs)', y=0.995)
	plt.tight_layout(rect=[0, 0, 1, 0.98])
	plt.show()

	return J_dict, age_jac_results