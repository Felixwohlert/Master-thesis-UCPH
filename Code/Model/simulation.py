"""
simulation.py — partial-equilibrium OLG wealth-distribution analysis.

Treats housing prices, rents, wages and the interest rate as EXOGENOUS (no
market clearing) and simulates the household block forward, recording the
net-wealth distribution by calendar time, age and housing tenure.

Public API:
    simulate_olg_pe(model, scenario)   — one PE simulation
    wealth_statistics(pe_result)       — Gini, percentile shares,
                                         homeownership, region / tenure shares
    build_scenario_grid(factors)       — Cartesian product of scenario factors
    build_scenario_ofat(factors)       — one-factor-at-a-time design (cheap)
    run_scenarios(model, scenarios)    — run many scenarios (serial)
    run_scenarios_parallel(...)        — same, dispatched across processes
    compare_scenarios(runs, stat)      — pull one statistic across runs
    plot_region_tenure(runs)           — 1x3 region / tenure panel
    plot_wealth_percentiles(runs)      — 2x2 net-wealth-share panel
    plot_region_wealth_stocks(runs, model)
                                       — 2x2 region-wise wealth decomposition
                                         across counterfactual scenarios
                                         (liquid A^u, A^r; housing q^u H^u,
                                         q^r H^r)
    plot_region_pq_decomposition(runs)
                                       — 3x2 housing wealth split into price
                                         q and demand H_hh channels (rows 1-2);
                                         product q·H_hh shown in row 3
"""

import itertools
import os

import numpy as np

import household_problem
from transition import (
    _get_path_cohort_weights,
    _get_hh_level_scale_path,
    _initialize_cohort_distribution,
    _scatter_dcegm_forward,
)


# ==========================================================================
# 1. Partial-equilibrium OLG simulation
# ==========================================================================

def simulate_olg_pe(model, scenario=None, n_nw_bins=400, nw_max=None):
    """Partial-equilibrium OLG simulation for wealth-distribution analysis.

    Treats the price paths on model.path as EXOGENOUS (no market clearing) and
    forward-simulates every cohort, recording the *net-wealth* distribution
    indexed by (calendar time, age, housing tenure).

    Net wealth of a household in state (z, h, a):
        net_wealth = a + housing_equity(h)
        housing_equity = (1 - lambda_ltv) * q^j * h^j   for owners (h=2,3)
                       = 0                               for renters (h=0,1)
    i.e. financial assets plus the un-mortgaged share of the house value.  All
    owners hold a mortgage at LTV lambda_ltv against q^j*h^j; T_mort is large
    so the principal is ~constant and the equity share is ~(1-lambda_ltv).

    Args:
        model:     HANCHousingModel with the steady state solved.
        scenario:  optional dict {name: length-T array} of exogenous paths
                   written onto model.path before simulating.  Price names
                   (r, w, q_u, q_r, f_u, f_r, kappa) and 'cohort_weights'
                   / 'hh_scale' are accepted.
        n_nw_bins: number of net-wealth histogram bins.
        nw_max:    top of the net-wealth grid (default: auto from prices/grid).

    Returns dict with:
        'nw_grid'  : (n_nw_bins,) net-wealth bin centres
        'nw_edges' : (n_nw_bins+1,) bin edges
        'nw_hist'  : (T, J, Nh, n_nw_bins) omega-weighted mass.  Summing over
                     (J, Nh) at a fixed t gives the population net-wealth
                     distribution at calendar time t.
        'omega'    : (T, J) cohort weights
        'A_hh','C_hh','H_u_hh','H_r_hh' : (T,) aggregates
        'D'        : (T, Nz, Nh, Na) aggregate distribution
        'T','J','Nh'
    """
    par = model.par
    path = model.path
    Nh = par.Nh if hasattr(par, 'Nh') else 4
    Nz = int(par.Nz)
    Na = int(par.Na)
    T = int(par.T)
    J = int(getattr(par, 'J', T))
    J = max(1, min(J, T))

    # --- Apply scenario paths -------------------------------------------------
    if scenario is not None:
        for name, vals in scenario.items():
            if name == 'cohort_weights':
                cw = np.asarray(vals, dtype=float)
                path.cohort_weights = (np.tile(cw[None, :], (T, 1))
                                       if cw.ndim == 1 else cw.copy())
                continue
            vec = np.asarray(vals, dtype=float).ravel()
            arr = path.__dict__.get(name)
            if isinstance(arr, np.ndarray) and arr.ndim == 2:
                arr[:, 0] = vec[:T]
            elif isinstance(arr, np.ndarray):
                arr[:T] = vec[:T]
            else:
                path.__dict__[name] = vec[:T].copy()

    omega_path = np.ascontiguousarray(
        _get_path_cohort_weights(model, J), dtype=np.float64)
    hh_scale = np.ascontiguousarray(
        _get_hh_level_scale_path(model), dtype=np.float64)
    D0_cohort, D_birth = _initialize_cohort_distribution(model, J, Nz, Nh, Na)
    D_birth = np.ascontiguousarray(D_birth, dtype=np.float64)

    # --- Grid / scalar constants ---------------------------------------------
    z_grid = np.ascontiguousarray(par.z_grid, dtype=np.float64)
    z_trans = np.ascontiguousarray(
        par.z_trans[0] if par.z_trans.ndim == 3 else par.z_trans, dtype=np.float64)
    m_grid = np.ascontiguousarray(par.m_grid, dtype=np.float64)
    cfloor = float(par.cfloor)
    zeta = float(par.zeta)
    zeta_renter = float(getattr(par, 'zeta_renter', 0.0))
    h_u = float(par.h_u)
    h_r = float(par.h_r)
    h_l = float(getattr(par, 'h_l', par.h_r))
    lambda_ltv = float(par.lambda_ltv)
    T_mort = float(par.T_mort)
    delta_H = float(par.delta_H)
    tau_wealth = float(getattr(par, 'tau_wealth', 0.0))
    tau_profits = float(getattr(par, 'tau_profits', 0.0))
    theta = float(par.theta)
    one_minus_ltv = 1.0 - lambda_ltv

    def _full(name):
        arr = path.__dict__.get(name)
        if isinstance(arr, np.ndarray) and arr.size > 0:
            vec = arr[:, 0] if arr.ndim == 2 else arr.ravel()[:T]
            return np.ascontiguousarray(vec, dtype=np.float64).copy()
        return None



    r_path = _full('r')
    w_path = _full('w')
    q_u_path = _full('q_u')
    q_r_path = _full('q_r')

    # --- Rents from the rental-sector FOC (mirrors blocks.rental_sector) -----
    # Competitive rental investors arbitrage:
    #   f^j_t = q^j_t (1 + r_t) - (1 - theta) q^j_{t+1} (1 - delta_H)
    # with the terminal lead  q^j_T = ss.q^j + rho_q_terminal * (q^j_{T-1} - ss.q^j).
    # Rents use the FORWARD price q_{t+1}, so a static-q fallback would be wrong
    # whenever q is moving.  Constructed here from the q / r scenario so the HH
    # block sees rents consistent with the prices; pass 'f_u'/'f_r' explicitly
    # in `scenario` to override.
    def _rental_rents(q_path, ss_q):
        rho_q = float(getattr(par, 'rho_q_terminal', 0.0))
        q_next = np.empty(T)
        q_next[:-1] = q_path[1:]
        q_next[-1] = ss_q + rho_q * (q_path[-1] - ss_q)
        return q_path * (1.0 + r_path) - (1.0 - theta) * q_next * (1.0 - delta_H)

    scen_has_fu = scenario is not None and 'f_u' in scenario
    scen_has_fr = scenario is not None and 'f_r' in scenario
    f_u_path = (_full('f_u') if scen_has_fu
                else _rental_rents(q_u_path, float(model.ss.q_u)))
    f_r_path = (_full('f_r') if scen_has_fr
                else _rental_rents(q_r_path, float(model.ss.q_r)))

    # Write the consistent rents back onto model.path so the DC-EGM solve below
    # and the forward simulation use the SAME f_u / f_r.
    for nm, vec in (('f_u', f_u_path), ('f_r', f_r_path)):
        arr = path.__dict__.get(nm)
        if isinstance(arr, np.ndarray) and arr.ndim == 2:
            arr[:, 0] = vec
        elif isinstance(arr, np.ndarray):
            arr[:T] = vec
        else:
            path.__dict__[nm] = vec.copy()

    chi_full = (np.ascontiguousarray(par.chi, dtype=np.float64)
                if hasattr(par, 'chi') else np.zeros(T))

    # --- DC-EGM policies on the (T, J) calendar-age grid ---------------------
    prices_dict = {}
    for name in ('r', 'w', 'q_u', 'q_r', 'f_u', 'f_r', 'kappa'):
        arr = path.__dict__.get(name)
        if isinstance(arr, np.ndarray) and arr.size > 0:
            prices_dict[name] = np.asarray(arr)
    hh = household_problem.HousingModel(model)
    a_arr, c_arr, pr_arr = hh.solve_dcegm_calendar_age_grid(prices_dict)
    a_arr = np.ascontiguousarray(a_arr, dtype=np.float64)
    c_arr = np.ascontiguousarray(c_arr, dtype=np.float64)
    pr_arr = np.ascontiguousarray(pr_arr, dtype=np.float64)

    # --- Net-wealth histogram grid -------------------------------------------
    if nw_max is None:
        eq_max = one_minus_ltv * max(float(np.max(q_u_path)) * h_u,
                                     float(np.max(q_r_path)) * h_r)
        nw_max = 1.05 * (float(np.max(m_grid)) + eq_max)
    nw_min = 0.0
    nw_edges = np.linspace(nw_min, nw_max, n_nw_bins + 1)
    nw_grid = 0.5 * (nw_edges[:-1] + nw_edges[1:])
    bin_w = (nw_max - nw_min) / n_nw_bins

    nw_hist = np.zeros((T, J, Nh, n_nw_bins))
    A_hh = np.zeros(T)
    C_hh = np.zeros(T)
    H_u_hh = np.zeros(T)
    H_r_hh = np.zeros(T)
    path_D_4d = np.zeros((T, Nz, Nh, Na))

    def _accumulate(D_age, cal_t, age):
        w_age = omega_path[cal_t, age]
        if w_age <= 0.0:
            return
        scale_t = hh_scale[cal_t]
        a_pol = a_arr[cal_t, age]
        c_pol = c_arr[cal_t, age]
        sw = scale_t * w_age
        A_hh[cal_t] += sw * np.sum(D_age * a_pol)
        C_hh[cal_t] += sw * np.sum(D_age * c_pol)
        H_u_hh[cal_t] += sw * (np.sum(D_age[:, 1, :]) * h_l
                               + np.sum(D_age[:, 3, :]) * h_u)
        H_r_hh[cal_t] += sw * (np.sum(D_age[:, 0, :]) * h_l
                               + np.sum(D_age[:, 2, :]) * h_r)
        path_D_4d[cal_t] += w_age * D_age
        # net wealth = financial assets (savings policy) + housing equity
        equity = np.zeros(Nh)
        if Nh >= 4:
            equity[2] = one_minus_ltv * q_r_path[cal_t] * h_r
            equity[3] = one_minus_ltv * q_u_path[cal_t] * h_u
        nw = a_pol + equity[None, :, None]            # (Nz, Nh, Na)
        idx = np.clip(((nw - nw_min) / bin_w).astype(np.int64),
                      0, n_nw_bins - 1)
        mass = w_age * D_age
        for ih in range(Nh):
            nw_hist[cal_t, age, ih] += np.bincount(
                idx[:, ih, :].ravel(),
                weights=mass[:, ih, :].ravel(),
                minlength=n_nw_bins)

    def _step(D_age, cal_t, age):
        D_next = np.zeros((Nz, Nh, Na))
        chi_age = float(chi_full[min(age + 1, T - 1)])
        _scatter_dcegm_forward(
            D_age, D_next,
            np.ascontiguousarray(a_arr[cal_t, age]),
            np.ascontiguousarray(pr_arr[cal_t, age]),
            z_trans, z_grid, m_grid,
            float(r_path[cal_t]), float(w_path[cal_t]) * np.exp(chi_age),
            float(q_u_path[cal_t]), float(q_r_path[cal_t]), cfloor,
            zeta, zeta_renter, h_u, h_r, h_l,
            float(f_u_path[cal_t]), float(f_r_path[cal_t]),
            lambda_ltv, T_mort, delta_H,
            tau_wealth, tau_profits,
        )
        s = np.sum(D_next)
        if s > 0.0 and np.isfinite(s):
            return D_next / s
        return D_birth.copy()

    # --- (1) Pre-transition cohorts: at cal_t=0 they are age j0 > 0 ----------
    for j0 in range(1, J):
        D_age = np.ascontiguousarray(D0_cohort[j0], dtype=np.float64).copy()
        max_cal_t = min(J - 1 - j0, T - 1)
        for cal_t in range(max_cal_t + 1):
            age = j0 + cal_t
            _accumulate(D_age, cal_t, age)
            if cal_t < max_cal_t:
                D_age = _step(D_age, cal_t, age)

    # --- (2) Post-transition cohorts: born at cal_t b in [0, T-1] ------------
    for birth in range(T):
        D_age = D_birth.copy()
        max_age = min(J - 1, T - 1 - birth)
        for age in range(max_age + 1):
            cal_t = birth + age
            _accumulate(D_age, cal_t, age)
            if age < max_age:
                D_age = _step(D_age, cal_t, age)

    return {
        'nw_grid': nw_grid,
        'nw_edges': nw_edges,
        'nw_hist': nw_hist,
        'omega': omega_path,
        'A_hh': A_hh, 'C_hh': C_hh, 'H_u_hh': H_u_hh, 'H_r_hh': H_r_hh,
        'D': path_D_4d,
        'T': T, 'J': J, 'Nh': Nh,
        # Price paths actually faced by this scenario — needed by
        # wealth_growth_decomposition to compute the returns component.
        'r_path': r_path, 'q_u_path': q_u_path, 'q_r_path': q_r_path,
        # Population scale path actually faced by this scenario — needed by
        # plot_region_wealth_stocks so the decomposition picks up each
        # scenario's own hh_scale (run_scenarios snapshot/restore otherwise
        # leaves model.path.hh_scale at the baseline after the run).
        'hh_scale': hh_scale,
    }


def wealth_statistics(pe_result, year0=1992):
    """Net-wealth distribution statistics from a simulate_olg_pe result.

    Per calendar time t, from the population net-wealth distribution
    (the (J, Nh)-summed histogram):
        gini, top1, top10, middle40, bottom50  — net-wealth concentration
        homeownership                          — owner mass / total
        mean_nw                                — mean net wealth
        mean_nw_urban, mean_nw_rural           — mean net wealth by region
        urban_pop_share, rural_pop_share       — population share by region
        urban_renter_share, rural_renter_share — renter share within a region
    and by age (T, J):
        gini_by_age, mean_nw_by_age, homeownership_by_age

    Housing states: 0 = rural_renter, 1 = urban_renter,
                    2 = rural_owner,  3 = urban_owner.
    Wealth shares are computed on the bin grid; accuracy improves with
    n_nw_bins.  top1 / top10 are top-percentile / top-decile shares;
    middle40 is the 50th-90th percentile share = 1 - top10 - bottom50.
    """
    nw_grid = pe_result['nw_grid']
    nw_hist = pe_result['nw_hist']           # (T, J, Nh, n_bins)
    T, J, Nh, n_bins = nw_hist.shape

    def _lorenz(mass):
        tot = mass.sum()
        if tot <= 0.0:
            return None
        m = mass / tot
        wealth = m * nw_grid
        W = wealth.sum()
        if W <= 0.0:
            return None
        return np.cumsum(m), np.cumsum(wealth) / W

    def _gini(mass):
        lz = _lorenz(mass)
        if lz is None:
            return np.nan
        cum_pop, cum_w = lz
        cp = np.concatenate([[0.0], cum_pop])
        cw = np.concatenate([[0.0], cum_w])
        area = np.sum((cw[1:] + cw[:-1]) * (cp[1:] - cp[:-1])) / 2.0
        return 1.0 - 2.0 * area

    
    def _shares(mass):
        """Return (top10, top1, bottom50) net-wealth shares.

        Percentile cut-offs are interpolated on the Lorenz curve, not snapped to
        an integer bin index — snapping makes the cut-off jump a whole bin as the
        distribution drifts, amplifying a small period-2 wiggle into a sawtooth.
        """
        lz = _lorenz(mass)
        if lz is None:
            return np.nan, np.nan, np.nan
        cum_pop, cum_w = lz
        cp = np.concatenate([[0.0], cum_pop])      # (0,0) anchor for interp
        cw = np.concatenate([[0.0], cum_w])
        return (1.0 - np.interp(0.9,  cp, cw),     # top10
                1.0 - np.interp(0.99, cp, cw),     # top1
                np.interp(0.5,  cp, cw))           # bottom50

    gini = np.full(T, np.nan)
    top1 = np.full(T, np.nan)
    top10 = np.full(T, np.nan)
    middle40 = np.full(T, np.nan)
    bottom50 = np.full(T, np.nan)
    homeown = np.full(T, np.nan)
    mean_nw = np.full(T, np.nan)
    mean_nw_urban = np.full(T, np.nan)
    mean_nw_rural = np.full(T, np.nan)
    urban_pop_share = np.full(T, np.nan)
    rural_pop_share = np.full(T, np.nan)
    urban_renter_share = np.full(T, np.nan)
    rural_renter_share = np.full(T, np.nan)
    gini_by_age = np.full((T, J), np.nan)
    mean_nw_by_age = np.full((T, J), np.nan)
    homeown_by_age = np.full((T, J), np.nan)

    for t in range(T):
        pop = nw_hist[t].sum(axis=(0, 1))            # (n_bins,)
        gini[t] = _gini(pop)
        t10, t1, b50 = _shares(pop)
        top10[t], top1[t], bottom50[t] = t10, t1, b50
        if np.isfinite(t10) and np.isfinite(b50):
            middle40[t] = max(0.0, 1.0 - t10 - b50)

        tot = pop.sum()
        if tot > 0.0:
            mean_nw[t] = np.sum(pop * nw_grid) / tot
            # mass by housing state: sum over ages (axis 0) and bins (axis 2)
            mass_h = nw_hist[t].sum(axis=(0, 2))     # (Nh,)
            m_rr, m_ur = mass_h[0], mass_h[1]
            m_ro, m_uo = mass_h[2], mass_h[3]
            urb = m_ur + m_uo
            rur = m_rr + m_ro
            urban_pop_share[t] = urb / tot
            rural_pop_share[t] = rur / tot
            homeown[t] = (m_ro + m_uo) / tot
            if urb > 0.0:
                urban_renter_share[t] = m_ur / urb
            if rur > 0.0:
                rural_renter_share[t] = m_rr / rur
            urb_d = (nw_hist[t, :, 1, :] + nw_hist[t, :, 3, :]).sum(axis=0)
            rur_d = (nw_hist[t, :, 0, :] + nw_hist[t, :, 2, :]).sum(axis=0)
            if urb_d.sum() > 0.0:
                mean_nw_urban[t] = np.sum(urb_d * nw_grid) / urb_d.sum()
            if rur_d.sum() > 0.0:
                mean_nw_rural[t] = np.sum(rur_d * nw_grid) / rur_d.sum()

        for j in range(J):
            popj = nw_hist[t, j].sum(axis=0)          # (n_bins,)
            tj = popj.sum()
            if tj > 0.0:
                gini_by_age[t, j] = _gini(popj)
                mean_nw_by_age[t, j] = np.sum(popj * nw_grid) / tj
                homeown_by_age[t, j] = (nw_hist[t, j, 2, :].sum()
                                        + nw_hist[t, j, 3, :].sum()) / tj

    return {
        'years': year0 + np.arange(T),
        'gini': gini,
        'top1': top1, 'top10': top10, 'middle40': middle40, 'bottom50': bottom50,
        'homeownership': homeown, 'mean_nw': mean_nw,
        'mean_nw_urban': mean_nw_urban, 'mean_nw_rural': mean_nw_rural,
        'urban_pop_share': urban_pop_share, 'rural_pop_share': rural_pop_share,
        'urban_renter_share': urban_renter_share,
        'rural_renter_share': rural_renter_share,
        'gini_by_age': gini_by_age, 'mean_nw_by_age': mean_nw_by_age,
        'homeownership_by_age': homeown_by_age,
    }


def wealth_growth_decomposition(pe_result, model, year0=1992, year_end=2024,
                                as_share=False, per_capita=True, do_print=True):
    """Decompose wealth growth into saving vs returns x assets vs housing.

    Builds the two baseline-decomposition tables over [year0, year_end]:
      table_5_1 : by net-wealth group  — Mean / Top 10% / Mid. 40% / Bot. 50%
      table_5_2 : by region x tenure   — Mean / Urban {region,renters,owners}
                                         / Rural {region,renters,owners}

    Method
    ------
    Groups are CROSS-SECTIONAL — a group is whoever is in it in each given
    year.  Per year t and group g, net wealth splits into financial assets A
    and housing equity HE, recovered from nw_hist: net wealth in bin b /
    housing state ih is nw_grid[b], of which the housing-equity part is the
    known constant (1-lambda)*q*h (0 for renters), so financial assets =
    nw_grid[b] - equity[ih, t].

    Returns are mechanical:
        returns -> assets  = sum_t  r_t * A_{g,t}
        returns -> housing = sum_t  (q_{t+1}/q_t - 1) * HE_{g,t}  (urban+rural)
    Saving is the residual stock change net of returns:
        saving  -> assets  = (A_{g,t1}  - A_{g,t0})  - returns->assets
        saving  -> housing = (HE_{g,t1} - HE_{g,t0}) - returns->housing
    With cross-sectional groups the saving residual also absorbs reranking /
    composition change — folded in, by the chosen convention.

    per_capita
    ----------
    If True (default), every group total A_g, HE_g is divided by the group
    head-count mass N_{g,t} before the decomposition, so the cells describe the
    *average group member* rather than the group aggregate.  This strips out
    the pure group-size effect: with cross-sectional region x tenure groups a
    shrinking rural population would otherwise drag rural A_g down even when
    each remaining rural household keeps saving.  For the fixed-size percentile
    groups of table 5.1 (always 10/40/50% of the population) N_{g,t} is
    constant, so per_capita is a no-op there — only table 5.2 changes.

    Each cell is reported as % of the group's own net wealth in `year0`.  The
    assets/housing split sums to the saving (resp. returns) subtotal, and
    saving + returns sums to the total % wealth growth.  With as_share=True the
    Mean row is subtracted from every group row, turning the group rows into the
    % growth of the group's wealth *share*.

    Returns
    -------
    {'table_5_1': {row: {...}}, 'table_5_2': {row: {...}}}, where each row dict
    holds, all in % of the group's year0 net wealth (or %-of-share if as_share):
        sav_A, sav_H  — saving contribution, split assets / housing
        ret_A, ret_H  — returns contribution, split assets / housing
        sav   = sav_A + sav_H   — total saving contribution
        ret   = ret_A + ret_H   — total returns contribution
        total = sav + ret       — total % wealth growth
    plus W0 (group net wealth in year0) and N0 / N1 (group head-count mass in
    year0 / year_end — the migration story behind table 5.2).
    """
    par = model.par

    for k in ('r_path', 'q_u_path', 'q_r_path'):
        if k not in pe_result:
            raise KeyError(
                f"pe_result is missing '{k}' — re-run simulate_olg_pe; the "
                "price paths were added to its output for this decomposition.")

    nw_grid = np.asarray(pe_result['nw_grid'], dtype=float)
    nw_hist = np.asarray(pe_result['nw_hist'], dtype=float)   # (T,J,Nh,n_bins)
    T, J, Nh, n_bins = nw_hist.shape

    r_path   = np.asarray(pe_result['r_path'],   dtype=float).ravel()
    q_u_path = np.asarray(pe_result['q_u_path'], dtype=float).ravel()
    q_r_path = np.asarray(pe_result['q_r_path'], dtype=float).ravel()

    t0 = 0
    t1 = int(year_end - year0)
    if not (0 < t1 < T):
        raise ValueError(f'year_end={year_end} is outside the simulated '
                          f'horizon [{year0}, {year0 + T - 1}].')

    one_minus_ltv = 1.0 - float(par.lambda_ltv)
    h_u, h_r = float(par.h_u), float(par.h_r)

    # mass per (year, housing-state, bin), summed over age
    m = nw_hist.sum(axis=1)                                   # (T, Nh, n_bins)

    # housing equity per (year, housing-state): renters 0, owners (1-ltv)*q*h
    equity = np.zeros((T, Nh))
    if Nh >= 4:
        equity[:, 2] = one_minus_ltv * q_r_path[:T] * h_r     # 2 = rural owner
        equity[:, 3] = one_minus_ltv * q_u_path[:T] * h_u     # 3 = urban owner

    # financial assets per (year, housing-state, bin) = net wealth - equity
    fin_assets = np.maximum(nw_grid[None, None, :] - equity[:, :, None], 0.0)

    def _decompose(weight):
        """weight: (T,Nh,n_bins) fraction of each cell's mass in the group."""
        sel = m * weight                                      # (T,Nh,n_bins)
        A   = (sel * fin_assets).sum(axis=(1, 2))             # (T,) aggregate
        HEu = ((sel[:, 3, :] * equity[:, 3, None]).sum(axis=1)
               if Nh >= 4 else np.zeros(T))
        HEr = ((sel[:, 2, :] * equity[:, 2, None]).sum(axis=1)
               if Nh >= 4 else np.zeros(T))
        N   = sel.sum(axis=(1, 2))                            # (T,) head-count mass
        N0, N1 = float(N[t0]), float(N[t1])

        if per_capita:
            # Per group member: divide aggregates by the group head-count so a
            # shrinking/growing group size (migration) no longer drives the
            # decomposition.  For fixed-size percentile groups N is constant
            # and this cancels in the % normalisation below.
            Nsafe = np.where(N > 1e-12, N, np.nan)
            A   = A   / Nsafe
            HEu = HEu / Nsafe
            HEr = HEr / Nsafe
        HE  = HEu + HEr
        W0  = float(A[t0] + HE[t0])

        gu = q_u_path[t0 + 1:t1 + 1] / np.maximum(q_u_path[t0:t1], 1e-12) - 1.0
        gr = q_r_path[t0 + 1:t1 + 1] / np.maximum(q_r_path[t0:t1], 1e-12) - 1.0
        ret_A = float(np.sum(r_path[t0:t1] * A[t0:t1]))
        ret_H = float(np.sum(gu * HEu[t0:t1] + gr * HEr[t0:t1]))
        sav_A = float((A[t1]  - A[t0])  - ret_A)
        sav_H = float((HE[t1] - HE[t0]) - ret_H)

        if not (W0 > 0.0) or not np.isfinite(W0):
            return dict(sav_A=np.nan, sav_H=np.nan, ret_A=np.nan, ret_H=np.nan,
                        sav=np.nan, ret=np.nan, total=np.nan,
                        W0=W0, N0=N0, N1=N1)
        f = 100.0 / W0
        sav_A, sav_H = sav_A * f, sav_H * f
        ret_A, ret_H = ret_A * f, ret_H * f
        return dict(sav_A=sav_A, sav_H=sav_H, ret_A=ret_A, ret_H=ret_H,
                    sav=sav_A + sav_H, ret=ret_A + ret_H,
                    total=sav_A + sav_H + ret_A + ret_H,
                    W0=W0, N0=N0, N1=N1)

    full = np.ones((T, Nh, n_bins))

    def _pct_weight(lo, hi):
        """(T,Nh,n_bins): fraction of each bin in net-wealth band (lo, hi]."""
        w = np.zeros((T, n_bins))
        for t in range(T):
            pop = m[t].sum(axis=0)
            tot = pop.sum()
            if tot <= 0.0:
                continue
            p = pop / tot
            cum = np.cumsum(p)
            cum_prev = cum - p
            overlap = np.minimum(cum, hi) - np.maximum(cum_prev, lo)
            w[t] = np.clip(overlap / np.maximum(p, 1e-300), 0.0, 1.0)
        return np.broadcast_to(w[:, None, :], (T, Nh, n_bins))

    def _state_weight(states):
        """(T,Nh,n_bins): 1 for the listed housing states, else 0."""
        w = np.zeros((T, Nh, n_bins))
        for ih in states:
            w[:, ih, :] = 1.0
        return w

    table_5_1 = {
        'Mean':     _decompose(full),
        'Top 10%':  _decompose(_pct_weight(0.90, 1.00)),
        'Mid. 40%': _decompose(_pct_weight(0.50, 0.90)),
        'Bot. 50%': _decompose(_pct_weight(0.00, 0.50)),
    }
    table_5_2 = {
        'Mean':           _decompose(full),
        'Urban region':   _decompose(_state_weight([1, 3])),
        'Urban -Renters': _decompose(_state_weight([1])),
        'Urban -Owners':  _decompose(_state_weight([3])),
        'Rural region':   _decompose(_state_weight([0, 2])),
        'Rural -Renters': _decompose(_state_weight([0])),
        'Rural -Owners':  _decompose(_state_weight([2])),
    }

    if as_share:
        _share_keys = ('sav_A', 'sav_H', 'ret_A', 'ret_H', 'sav', 'ret', 'total')

        def _to_share(tbl):
            mean = tbl['Mean']
            out = {'Mean': mean}
            for name, d in tbl.items():
                if name == 'Mean':
                    continue
                out[name] = {k: (d[k] - mean[k] if k in _share_keys else d[k])
                             for k in d}
            return out
        table_5_1 = _to_share(table_5_1)
        table_5_2 = _to_share(table_5_2)

    if do_print:
        kind = 'wealth-SHARE growth' if as_share else 'wealth growth'
        note = ', per capita' if per_capita else ''

        def _print(title, tbl, labels):
            print(f'\n  {title}')
            print(f'  ({year0}-{year_end}, {kind}{note}, % of {year0} net wealth)')
            print(f'  {"":18s}{"%-growth":>15s}{"":20s}{"%-growth":>16s}')
            print(f'  {"group":18s}{"due to saving":>15s}{"Assets":>10s}'
                  f'{"Housing":>10s}{"due to returns":>16s}{"Assets":>10s}'
                  f'{"Housing":>10s}{"Total":>10s}')
            print('  ' + '-' * 97)
            for disp, key in labels:
                d = tbl[key]
                print(f'  {disp:18s}{d["sav"]:15.2f}{d["sav_A"]:10.2f}'
                      f'{d["sav_H"]:10.2f}{d["ret"]:16.2f}{d["ret_A"]:10.2f}'
                      f'{d["ret_H"]:10.2f}{d["total"]:10.2f}')

        _print('Table 5.1 - by net-wealth group', table_5_1, [
            ('Mean', 'Mean'), ('Top 10%', 'Top 10%'),
            ('Mid. 40%', 'Mid. 40%'), ('Bot. 50%', 'Bot. 50%')])
        _print('Table 5.2 - by region x tenure', table_5_2, [
            ('Mean', 'Mean'),
            ('Urban region', 'Urban region'),
            ('- Renters', 'Urban -Renters'), ('- Owners', 'Urban -Owners'),
            ('Rural region', 'Rural region'),
            ('- Renters', 'Rural -Renters'), ('- Owners', 'Rural -Owners')])

    return {'table_5_1': table_5_1, 'table_5_2': table_5_2}


def wealth_growth_decomposition_table(
        model, scenarios, runs=None,
        year0=1992, year_end=2024, per_capita=True,
        rows=None, col_labels=None, do_print=True, latex=False):
    """Cross-scenario wealth-growth decomposition: liquid (A) vs housing (H).

    Builds a table whose COLUMNS are scenarios and whose ROWS are region x
    tenure groups.  For every (scenario, group) cell the % wealth growth from
    `year0` to `year_end` is split into two contributions that sum to the
    group's total % growth:

        A = sav_A + ret_A   (liquid assets: saving + returns)
        H = sav_H + ret_H   (housing:       saving + returns)
        A + H = total % wealth growth

    The per-cell numbers come straight from `wealth_growth_decomposition`
    (its `table_5_2`), run once per scenario on THAT scenario's own model, so
    policy parameters (lambda_ltv, theta, the taxes) enter the equity and rent
    accounting correctly.

    Parameters
    ----------
    model : default HANCHousingModel (used for scenarios without a 'model' key).
    scenarios : dict {name: {... ['model': m] ...}} — the same object passed to
        `run_scenarios`.  Column order follows this dict's insertion order.
    runs : optional output of `run_scenarios`.  When given, runs[name]['pe'] is
        reused (no re-simulation).  When None, the scenarios are simulated here
        via `run_scenarios`.
    rows : optional list of (display_label, table_5_2_key).  Default is
        Mean / Urban renters / Urban owners / Rural renters / Rural owners.
    col_labels : optional {name: header} to relabel columns for display/LaTeX,
        e.g. {'profit tax': r'Profit tax, $\\tau^q$'}.
    latex : if True, also print a paste-ready LaTeX tabular.

    Returns
    -------
    dict {scenario_name: {display_label: {'A':.., 'H':.., 'total':..}}}
    """
    if rows is None:
        rows = [
            ('Mean',          'Mean'),
            ('Urban renters', 'Urban -Renters'),
            ('Urban owners',  'Urban -Owners'),
            ('Rural renters', 'Rural -Renters'),
            ('Rural owners',  'Rural -Owners'),
        ]
    col_labels = col_labels or {}
    names = list(scenarios.keys())

    if runs is None:
        runs = run_scenarios(model, scenarios, include_baseline=False,
                             year0=year0, do_print=do_print)

    out = {}
    for name in names:
        scen_clean, scen_model = _split_model_from_scenario(scenarios[name], model)
        pe = None
        if isinstance(runs, dict) and isinstance(runs.get(name), dict):
            pe = runs[name].get('pe')
        if pe is None:
            raise KeyError(
                f"No 'pe' result found for scenario '{name}'. Pass runs=None to "
                "simulate here, or supply a `runs` dict from run_scenarios that "
                "contains it.")
        dec = wealth_growth_decomposition(
            pe, scen_model, year0=year0, year_end=year_end,
            per_capita=per_capita, do_print=False)
        t2 = dec['table_5_2']
        col = {}
        for disp, key in rows:
            d = t2.get(key, {})
            A = float(d.get('sav_A', np.nan)) + float(d.get('ret_A', np.nan))
            H = float(d.get('sav_H', np.nan)) + float(d.get('ret_H', np.nan))
            col[disp] = {'A': A, 'H': H, 'total': A + H}
        out[name] = col

    def _hdr(n):
        return col_labels.get(n, n)

    if do_print:
        def _c(x):
            return f'{x:8.1f}' if np.isfinite(x) else '     nan'
        title = (f'\nWealth-growth decomposition by scenario '
                 f'({year0}-{year_end}, % of {year0} net wealth; '
                 f'A+H = total % growth'
                 f'{", per capita" if per_capita else ""})')
        print(title)
        line1 = f'{"":18s}'
        for n in names:
            line1 += f'{_hdr(n)[:15]:^16s}'
        line2 = f'{"":18s}' + ''.join(f'{"A":>8s}{"H":>8s}' for _ in names)
        print(line1)
        print(line2)
        print('-' * (18 + 16 * len(names)))
        for disp, _key in rows:
            line = f'{disp:18s}'
            for n in names:
                c = out[n][disp]
                line += f'{_c(c["A"])}{_c(c["H"])}'
            print(line)

    if latex:
        print(_latex_decomposition_table(out, names, rows, col_labels,
                                         year0, year_end, per_capita))
    return out


_SCENARIO_PATH_NAMES = ('r', 'w', 'q_u', 'q_r', 'f_u', 'f_r', 'kappa',
                        'cohort_weights', 'hh_scale')


def _split_model_from_scenario(scen, default_model):
    """Pull an optional 'model' key out of a scenario dict.

    Returns (scen_clean, model_for_run). 'model' is a reserved meta-key that
    selects a different HANCHousingModel for this scenario (useful for varying
    structural parameters — e.g. par.beta — that aren't path inputs to
    simulate_olg_pe). The path-patch dict actually fed to simulate_olg_pe has
    'model' stripped; if no path patches remain it becomes None.
    """
    if scen is None or 'model' not in scen:
        return scen, default_model
    m = scen['model']
    rest = {k: v for k, v in scen.items() if k != 'model'}
    return (rest if rest else None), m


def _build_scenario_todo(model, scenarios, include_baseline, baseline_name):
    """Resolve scenarios into a list of (name, scen_clean, model_for_run) tuples.

    Used by both serial and parallel run_scenarios variants so the dispatch
    rules — including the 'model' meta-key — stay consistent.
    """
    todo = []
    if include_baseline:
        todo.append((baseline_name, None, model))
    for name, scen in scenarios.items():
        scen_clean, scen_model = _split_model_from_scenario(scen, model)
        todo.append((name, scen_clean, scen_model))
    return todo


def run_scenarios(model, scenarios, include_baseline=True,
                  baseline_name='baseline', n_nw_bins=400, nw_max=None,
                  year0=1992, do_print=True):
    """Run simulate_olg_pe + wealth_statistics across a set of named scenarios.

    A *scenario* is a dict {varname: length-T path} that PATCHES the baseline
    paths currently on model.path; variables not listed keep their baseline
    values.  Accepted varnames: r, w, q_u, q_r, f_u, f_r, kappa,
    cohort_weights, hh_scale — i.e. the exogenous price/preference/demographic
    inputs to the household block.

    Per-scenario model override
    ---------------------------
    A scenario may additionally carry the reserved key 'model', whose value is
    a fully-solved HANCHousingModel instance.  That scenario will run on the
    supplied model instead of the default `model` argument, which lets you
    vary structural parameters (e.g. par.beta, par.rho, par.zeta) that are NOT
    path inputs. The override may be combined with path patches:

        scenarios = {
            'baseline':            {'q_u': q_u_base},
            'high_beta':           {'model': model_high_beta},
            'high_beta_low_r':     {'model': model_high_beta, 'r': r_low_path},
        }

    The 'model' key is stripped before scenarios are passed to simulate_olg_pe.
    The baseline path of every distinct model that will be touched is snapshotted
    up front and restored before each run, so scenarios stay independent.

    A common net-wealth grid is used across all scenarios so their histograms
    and statistics are directly comparable.

    Args:
        model:        Default HANCHousingModel, SS solved, baseline paths on
                      model.path. Used by scenarios that don't supply a 'model'.
        scenarios:    dict {scenario_name: {varname: path, ..., ['model': m]}}.
        include_baseline: also run the unpatched baseline under baseline_name.
        n_nw_bins:    net-wealth histogram resolution.
        nw_max:       net-wealth grid ceiling; default = common auto ceiling
                      scanning all scenarios' q paths AND all per-scenario
                      models so the grid is shared.
        year0:        calendar year of t=0.

    Returns:
        dict {scenario_name: {'pe': <simulate_olg_pe result>,
                              'stats': <wealth_statistics result>}}
    """
    todo = _build_scenario_todo(model, scenarios, include_baseline, baseline_name)

    # Snapshot the baseline path of every distinct model that any scenario uses.
    # Keyed by id() to share snapshots when scenarios reuse the same instance.
    snapshots = {}                          # id(m) -> (m, {name: array})
    for _, _, m in todo:
        if id(m) in snapshots:
            continue
        snap = {}
        for nm in _SCENARIO_PATH_NAMES:
            arr = m.path.__dict__.get(nm)
            if isinstance(arr, np.ndarray):
                snap[nm] = arr.copy()
        snapshots[id(m)] = (m, snap)

    def _restore_all():
        for m_obj, snap in snapshots.values():
            for nm, arr in snap.items():
                cur = m_obj.path.__dict__.get(nm)
                if isinstance(cur, np.ndarray) and cur.shape == arr.shape:
                    cur[:] = arr
                else:
                    m_obj.path.__dict__[nm] = arr.copy()

    # Common net-wealth ceiling across all (scenario, model) pairs.
    if nw_max is None:
        def _scen_q(scen, m_obj, name):
            base = snapshots[id(m_obj)][1].get(name)
            if base is not None:
                base = base[:, 0] if base.ndim == 2 else base.ravel()
            if scen is not None and name in scen:
                return np.asarray(scen[name], dtype=float).ravel()
            return base if base is not None else np.array([0.0])

        nw_max = 0.0
        for _, scen, m_obj in todo:
            par_m = m_obj.par
            one_minus_ltv = 1.0 - float(par_m.lambda_ltv)
            h_u_m  = float(par_m.h_u)
            h_r_m  = float(par_m.h_r)
            mmax_m = float(np.max(par_m.m_grid))
            qu = _scen_q(scen, m_obj, 'q_u')
            qr = _scen_q(scen, m_obj, 'q_r')
            scen_ceiling = 1.05 * (
                mmax_m + one_minus_ltv * max(float(np.max(qu)) * h_u_m,
                                             float(np.max(qr)) * h_r_m)
            )
            nw_max = max(nw_max, scen_ceiling)

    runs = {}
    try:
        for name, scen, m_obj in todo:
            if do_print:
                marker = '' if m_obj is model else '  [alt model]'
                print(f'  scenario: {name}{marker} ...', flush=True)
            _restore_all()
            pe = simulate_olg_pe(m_obj, scenario=scen,
                                 n_nw_bins=n_nw_bins, nw_max=nw_max)
            stats = wealth_statistics(pe, year0=year0)
            runs[name] = {'pe': pe, 'stats': stats}
    finally:
        _restore_all()

    return runs


def _resolve_nw_max(todo):
    """Shared net-wealth ceiling across all scenarios (common histogram grid).

    todo: list of (name, scen, model) tuples as produced by
    _build_scenario_todo. Each scenario's ceiling uses its own model's housing
    parameters and m_grid, so the shared grid is large enough for every run.
    Reads the current model.path values as the baseline (run_scenarios_parallel
    never mutates model.path, so no snapshot is needed).
    """
    def _base(m_obj, nm):
        a = m_obj.path.__dict__.get(nm)
        if isinstance(a, np.ndarray):
            return a[:, 0] if a.ndim == 2 else a.ravel()
        return np.array([0.0])

    nw_max = 0.0
    for _, scen, m_obj in todo:
        par_m = m_obj.par
        one_minus_ltv = 1.0 - float(par_m.lambda_ltv)
        h_u  = float(par_m.h_u)
        h_r  = float(par_m.h_r)
        mmax = float(np.max(par_m.m_grid))
        qu = (np.asarray(scen['q_u'], dtype=float).ravel()
              if (scen and 'q_u' in scen) else _base(m_obj, 'q_u'))
        qr = (np.asarray(scen['q_r'], dtype=float).ravel()
              if (scen and 'q_r' in scen) else _base(m_obj, 'q_r'))
        scen_ceiling = 1.05 * (
            mmax + one_minus_ltv * max(float(np.max(qu)) * h_u,
                                       float(np.max(qr)) * h_r)
        )
        nw_max = max(nw_max, scen_ceiling)
    return nw_max


def _run_one_scenario(model, name, scen, nw_max, n_nw_bins, year0):
    """Worker entry point: one scenario → (name, {'pe', 'stats'}).

    Module-level (not a closure) so joblib can pickle it.  Each worker gets
    its own cloudpickled copy of `model`, so simulate_olg_pe's in-place
    patching of model.path cannot leak between scenarios.
    """
    pe = simulate_olg_pe(model, scenario=scen,
                         n_nw_bins=n_nw_bins, nw_max=nw_max)
    stats = wealth_statistics(pe, year0=year0)
    return name, {'pe': pe, 'stats': stats}


def run_scenarios_parallel(model, scenarios, include_baseline=True,
                           baseline_name='baseline', n_nw_bins=400, nw_max=None,
                           year0=1992, n_workers=None, do_print=True):
    """Parallel run_scenarios — independent scenarios dispatched across processes.

    Same arguments, same return value, and the same per-scenario 'model' meta-
    key convention as run_scenarios (see that docstring for details).  Each
    scenario is a full, independent OLG simulation (no work is shared between
    them), so the set parallelises perfectly.  joblib's loky backend
    cloudpickles whichever model the scenario selects to each worker; the
    caller's model(s) are never mutated, and loky caps native-library threads
    per worker so the processes do not oversubscribe the cores.

    Args:
        model, scenarios, include_baseline, baseline_name, n_nw_bins, nw_max,
        year0, do_print:  as run_scenarios.
        n_workers:  process count (default min(#scenarios, os.cpu_count())).

    Returns:
        dict {scenario_name: {'pe': ..., 'stats': ...}}.
    """
    try:
        from joblib import Parallel, delayed
    except ImportError as exc:
        raise ImportError(
            "run_scenarios_parallel requires joblib (`pip install joblib`); "
            "use run_scenarios for the serial version.") from exc

    todo = _build_scenario_todo(model, scenarios, include_baseline, baseline_name)

    # Resolve the shared net-wealth grid once, in the parent, so every
    # scenario's histogram is directly comparable.
    if nw_max is None:
        nw_max = _resolve_nw_max(todo)

    if n_workers is None:
        n_workers = min(len(todo), os.cpu_count() or 1)
    n_workers = max(1, min(int(n_workers), len(todo)))

    if do_print:
        n_alt = sum(1 for _, _, m_ in todo if m_ is not model)
        alt_note = f' ({n_alt} on alt model)' if n_alt else ''
        print(f'  run_scenarios_parallel: {len(todo)} scenarios{alt_note} on '
              f'{n_workers} worker process(es) ...', flush=True)

    results = Parallel(n_jobs=n_workers, backend='loky',
                       verbose=10 if do_print else 0)(
        delayed(_run_one_scenario)(m_obj, name, scen,
                                   nw_max, n_nw_bins, year0)
        for name, scen, m_obj in todo)

    return {name: res for name, res in results}


def compare_scenarios(runs, stat='gini'):
    """Pull one wealth statistic across scenarios for plotting/tabulation.

    Args:
        runs: output of run_scenarios.
        stat: key in the wealth_statistics result — e.g. 'gini', 'top10',
              'top1', 'bottom50', 'homeownership', 'mean_nw',
              'mean_nw_urban', 'mean_nw_rural'.

    Returns (years, {scenario_name: stat_array}).
    """
    years = None
    series = {}
    for name, r in runs.items():
        st = r['stats']
        if years is None:
            years = st['years']
        series[name] = st[stat]
    return years, series


def build_scenario_grid(factors):
    """Cartesian product of scenario factors.

    A *factor* is one exogenous driver you want to vary (e.g. the interest
    rate, urban preference).  Each factor offers a set of named *options*,
    and each option is a dict of path-patches.

    Args:
        factors: dict {factor_name: {option_name: {varname: path}}}, e.g.
            {
              'r':     {'low':  {'r': r_low},
                        'base': {'r': r_base},
                        'high': {'r': r_high}},
              'kappa': {'flat':   {'kappa': k_flat},
                        'rising': {'kappa': k_rising}},
            }

    Returns:
        dict {combo_label: merged_scenario_dict} containing every combination
        of one option per factor — combo_label e.g. 'r=low | kappa=flat'.
        Feed the result straight to run_scenarios.
    """
    factor_names = list(factors.keys())
    option_lists = [list(factors[fn].items()) for fn in factor_names]

    grid = {}
    for combo in itertools.product(*option_lists):
        label_parts = []
        merged = {}
        for fname, (oname, scen) in zip(factor_names, combo):
            label_parts.append(f"{fname}={oname}")
            merged.update(scen)
        grid[" | ".join(label_parts)] = merged
    return grid


def build_scenario_ofat(factors, baseline=None, baseline_label='baseline'):
    """One-factor-at-a-time (OFAT) scenario design.

    A cheaper alternative to build_scenario_grid's full Cartesian product.
    build_scenario_grid produces ∏_f n_options(f) scenarios (e.g. five
    factors with 3,3,3,2,2 options → 108); this produces only

        1 + Σ_f (n_options(f) − 1)

    scenarios — one all-baseline run plus, for each factor, one run per
    off-baseline option with every other factor held at baseline.  For the
    same five factors that is 1 + (2+2+2+1+1) = 9.

    Args:
        factors: same structure as build_scenario_grid,
            {factor_name: {option_name: {varname: path}}}.
        baseline: optional {factor_name: option_name} naming the baseline
            option of each factor.  Any factor not listed is auto-detected:
            the first option named 'base' / 'baseline' / 'mid' if present,
            else the first option defined for that factor.
        baseline_label: label of the all-baseline run.

    Returns:
        dict {label: merged_scenario_dict} — baseline_label for the
        all-baseline run and one f"{factor}={option}" entry per off-baseline
        option.  The baseline is already included, so feed the result to
        run_scenarios / run_scenarios_parallel with include_baseline=False.
    """
    factor_names = list(factors.keys())
    baseline = dict(baseline) if baseline else {}

    # Resolve each factor's baseline option.
    base_opt = {}
    for fn in factor_names:
        opts = list(factors[fn])
        if fn in baseline:
            if baseline[fn] not in factors[fn]:
                raise ValueError(
                    f"baseline['{fn}'] = '{baseline[fn]}' is not an option of "
                    f"factor '{fn}' (options: {opts})")
            base_opt[fn] = baseline[fn]
        else:
            base_opt[fn] = next((c for c in ('base', 'baseline', 'mid')
                                 if c in factors[fn]), opts[0])

    def _all_baseline():
        merged = {}
        for fn in factor_names:
            merged.update(factors[fn][base_opt[fn]])
        return merged

    grid = {baseline_label: _all_baseline()}
    for fn in factor_names:
        for oname, patch in factors[fn].items():
            if oname == base_opt[fn]:
                continue
            scen = _all_baseline()
            scen.update(patch)
            grid[f"{fn}={oname}"] = scen
    return grid


# ==========================================================================
# 5. Plotting
# ==========================================================================

def plot_region_tenure(runs, figsize=(15, 4), xlim=None,
                       baseline_name='baseline',
                       style='auto', lines_threshold=5):
    """1x3 panel: urban population share, urban renter share, rural renter share.

    Plotting style:
      - The scenario named `baseline_name` (default 'baseline') is drawn as
        a solid blue line.
      - Every other scenario in `runs` is shown either as per-scenario named
        lines or bundled into a grey min-max fan (see `style`).
      - The empirical series from datagraphs.xlsx is overlaid as a dashed
        black `data` line.
    Pass baseline_name=None to put every scenario in the secondary group.

    Args:
        style           : 'auto' (default) draws individual coloured lines when
                          the number of non-baseline scenarios is <=
                          lines_threshold, otherwise a min-max fan. 'lines'
                          forces per-scenario lines; 'fan' forces the fan.
        lines_threshold : cutoff used by style='auto' (default 5).

    Returns the matplotlib Figure.
    """
    import matplotlib.pyplot as plt

    import openpyxl as _openpyxl
    _wb = _openpyxl.load_workbook('datagraphs.xlsx', data_only=True)
    _ws = _wb.active
    _headers = [c.value for c in _ws[1]]

    # datagraphs.xlsx is inconsistent about casing ('Urban Region' vs
    # 'Rural region'), so match column names case-insensitively.
    _headers_lower = [str(h).lower() if h is not None else None for h in _headers]
    def _col(name):
        return _headers_lower.index(name.lower())

    urb_renter_share = _col('Share of renters - Urban Region')
    rural_renter_share = _col('Share of renters - Rural Region')
    urban_pop_share = _col('Urban population share')

    _urb_renter_share = np.array(
        [row[urb_renter_share] for row in _ws.iter_rows(min_row=2, values_only=True) if row[urb_renter_share] is not None],
        dtype=float,
        ) 
    
    _rural_renter_share = np.array(
        [row[rural_renter_share] for row in _ws.iter_rows(min_row=2, values_only=True) if row[rural_renter_share] is not None],
        dtype=float,
    )   

    _urban_pop_share = np.array(
        [row[urban_pop_share] for row in _ws.iter_rows(min_row=2, values_only=True) if row[urban_pop_share] is not None],
        dtype=float,
    )

    # Empirical series, keyed to match the panel keys below.
    emp = {
        'urban_pop_share':    _urban_pop_share,
        'urban_renter_share': _urb_renter_share,
        'rural_renter_share': _rural_renter_share,
    }

    panels = [
        ('urban_pop_share',    'Urban population share'),
        ('urban_renter_share', 'Renter share - urban region'),
        ('rural_renter_share', 'Renter share - rural region'),
    ]
    # Separate the baseline (highlighted) from the other scenarios (fan).
    baseline = runs.get(baseline_name) if baseline_name is not None else None
    fan_names = [n for n in runs if n != baseline_name]
    years_ref = (baseline['stats']['years'] if baseline is not None
                 else (next(iter(runs.values()))['stats']['years']
                       if runs else np.array([])))

    use_lines = _resolve_scenario_style(style, fan_names, lines_threshold)

    fig, axes = plt.subplots(1, 3, figsize=(9, 5))
    for ax, (key, title) in zip(axes, panels):
        if fan_names:
            if use_lines:
                # per-scenario coloured lines (skip C0 — reserved for baseline)
                for i, n in enumerate(fan_names):
                    ax.plot(years_ref, runs[n]['stats'][key],
                            lw=1.8, color=f'C{i + 1}', alpha=0.9, label=n)
            else:
                # grey min-max fan across the non-baseline scenarios
                stack = np.stack([runs[n]['stats'][key] for n in fan_names])
                ax.fill_between(years_ref, stack.min(axis=0), stack.max(axis=0),
                                color='0.7', alpha=0.4, linewidth=0,
                                label=f'scenarios ({len(fan_names)})')
        # baseline line
        col_baseline = '#003C8F'   # dark blue
        if baseline is not None:
            ax.plot(years_ref, baseline['stats'][key],
                    color=col_baseline, lw=2, label=baseline_name)
        # empirical
        emp_arr = emp[key]
        emp_years = 1992 + np.arange(len(emp_arr))   # xlsx starts at 1992
        ax.plot(emp_years, emp_arr, 'k--', lw=2, label='data')
        ax.set_title(title)
        ax.set_xlabel('year')
        if xlim is not None:
            ax.set_xlim(*xlim)
        ax.set_ylim(0, 1.0)
        ax.set_box_aspect(1.0)
    axes[0].legend(fontsize=8, loc='best', frameon=False)
    fig.suptitle('Region and tenure shares')
    fig.tight_layout()
    return fig


def _resolve_scenario_style(style, fan_names, lines_threshold):
    """Decide whether the non-baseline scenarios get individual lines or a fan.

    Returns True for per-scenario lines, False for the min-max fan. Used by
    both plot_region_tenure and plot_wealth_percentiles so they treat 'style'
    identically.
    """
    s = str(style).lower()
    if s == 'lines':
        return True
    if s == 'fan':
        return False
    if s == 'auto':
        return len(fan_names) <= int(lines_threshold)
    raise ValueError(f"style must be 'auto', 'lines' or 'fan' (got {style!r})")


def plot_wealth_percentiles(runs, figsize=(11, 7), xlim=None, ma=0,
                            baseline_name='baseline',
                            style='auto', lines_threshold=5):
    """2x2 panel: top-1%, top-10%, middle-40%, bottom-50% net-wealth shares.

    Plotting style:
      - The scenario named `baseline_name` (default 'baseline') is drawn as
        a solid blue line.
      - Every other scenario in `runs` is shown either as per-scenario named
        lines or bundled into a grey min-max fan (see `style`).
      - The empirical series from datagraphs.xlsx is overlaid as a dashed
        black `data` line.
    Pass baseline_name=None to put every scenario in the secondary group.

    Returns the matplotlib Figure.

    ma : int
        Moving-average window applied to every plotted series (baseline, fan
        envelope, and data).  ma=0 (or 1) plots raw series; ma>=2 replaces
        them with a centered moving average (e.g. ma=4, ma=5).
    style : str
        'auto' (default) — coloured lines per scenario when there are <=
        lines_threshold non-baseline scenarios, otherwise the min-max fan.
        'lines' forces individual lines; 'fan' forces the grey envelope.
    lines_threshold : int
        Cutoff used by style='auto' (default 5).
    """
    import matplotlib.pyplot as plt
    import openpyxl as _openpyxl
    _wb = _openpyxl.load_workbook('datagraphs.xlsx', data_only=True)
    _ws = _wb.active
    _headers = [c.value for c in _ws[1]]
    _top_1 = _headers.index('Top 1')
    _top_10 = _headers.index('Top 10')
    _middle_40 = _headers.index('Middle 40')
    _bottom_50 = _headers.index('Bottom 50')


    # b. Extracting the annual data and transforming appropriately
    top_1 = np.array(
        [row[_top_1] for row in _ws.iter_rows(min_row=2, values_only=True) if row[_top_1] is not None],
        dtype=float,
        ) 

    top_10 = np.array(
        [row[_top_10] for row in _ws.iter_rows(min_row=2, values_only=True) if row[_top_10] is not None],
        dtype=float,
    ) 

    middle_40 = np.array(
        [row[_middle_40] for row in _ws.iter_rows(min_row=2, values_only=True) if row[_middle_40] is not None],
        dtype=float,
    ) 

    bottom_50 = np.array(
        [row[_bottom_50] for row in _ws.iter_rows(min_row=2, values_only=True) if row[_bottom_50] is not None],
        dtype=float,
    ) 

    emp = {'top1': top_1, 'top10': top_10,
           'middle40': middle_40, 'bottom50': bottom_50}
    emp_years = 1992 + np.arange(len(top_1))   # xlsx is 1992..2024, ascending

    panels = [
        ('top1',     'Top 1% net-wealth share'),
        ('top10',    'Top 10% net-wealth share'),
        ('middle40', 'Middle 40% net-wealth share'),
        ('bottom50', 'Bottom 50% net-wealth share'),
    ]
    def _moving_average(y, window):
        """Centered moving average; window<=1 returns y unchanged.

        Edge points use a shrinking (partial) window so the smoothed
        series keeps the same length as the input.
        """
        y = np.asarray(y, dtype=float)
        n = y.size
        if window <= 1 or n == 0:
            return y
        half_lo = window // 2
        half_hi = window - 1 - half_lo
        out = np.empty(n)
        for i in range(n):
            lo = max(0, i - half_lo)
            hi = min(n, i + half_hi + 1)
            out[i] = np.mean(y[lo:hi])
        return out

    # Separate the baseline (highlighted) from the other scenarios (fan).
    baseline = runs.get(baseline_name) if baseline_name is not None else None
    fan_names = [n for n in runs if n != baseline_name]
    years_ref = (baseline['stats']['years'] if baseline is not None
                 else (next(iter(runs.values()))['stats']['years']
                       if runs else np.array([])))

    use_lines = _resolve_scenario_style(style, fan_names, lines_threshold)

    fig, axes = plt.subplots(2, 2, figsize=(8,8))
    for ax, (key, title) in zip(axes.ravel(), panels):
        if fan_names:
            if use_lines:
                # per-scenario coloured lines (skip C0 — reserved for baseline)
                for i, n in enumerate(fan_names):
                    ax.plot(years_ref,
                            _moving_average(runs[n]['stats'][key], ma),
                            lw=1.8, color=f'C{i + 1}', alpha=0.9, label=n)
            else:
                # grey min-max fan across the non-baseline scenarios
                stack = np.stack([_moving_average(runs[n]['stats'][key], ma)
                                  for n in fan_names])
                ax.fill_between(years_ref, stack.min(axis=0), stack.max(axis=0),
                                color='0.7', alpha=0.4, linewidth=0,
                                label=f'scenarios ({len(fan_names)})')
        # baseline line
        col_baseline = '#003C8F'   # dark blue
        if baseline is not None:
            ax.plot(years_ref, _moving_average(baseline['stats'][key], ma),
                    color=col_baseline, lw=2, label=baseline_name)
        # empirical
        ax.plot(emp_years, _moving_average(emp[key], ma),
                'k--', lw=2, label='data')
        ax.set_title(title)
        ax.set_xlabel('year')
        if xlim is not None:
            ax.set_xlim(*xlim)
        ax.set_ylim(0, 0.75)
        ax.set_box_aspect(1.0)
    axes.ravel()[0].legend(fontsize=8, loc='best', frameon=False)

    suptitle = 'Net-wealth shares by percentile group'
    if ma > 1:
        suptitle += f'  ({ma}-period moving average)'
    fig.suptitle(suptitle)
    fig.tight_layout()
    return fig


def plot_region_wealth_stocks(runs, model, year0=1992, year_end=2024,
                              baseline_name='baseline', normalize='level',
                              colors=None, legend_labels=None,
                              legend_loc='best', legend_bbox=None,
                              data_file='datagraphs.xlsx', show_data=True,
                              figsize=(13, 9), title=None):
    """2x2 decomposition of region-wise wealth stocks across counterfactuals.

    Layout (each panel plots the CHANGE in wealth from year0):
      Top-left    : Urban liquid wealth   A^u
      Top-right   : Rural liquid wealth   A^r
      Bottom-left : Urban housing wealth  q^u * H^u
      Bottom-right: Rural housing wealth  q^r * H^r

    Per panel:
      - The cumulative change of the baseline (dashed black line).
      - For every non-baseline scenario in `runs`, its CONTRIBUTION
            contribution_X(t) = baseline_level(t) - scenario_X_level(t)
        is plotted as a filled layer.  Positive contributions stack ABOVE
        the zero line, negative ones stack BELOW it — so each scenario uses
        a single color and may appear on both sides of zero over time
        depending on its sign.  Each contribution is what the driver that
        scenario_X *fixes* would have added by being left to transition.

    Caveat: with one-at-a-time counterfactuals the contributions do NOT
    generally sum to the baseline change because the model is nonlinear
    (Shapley-style additivity would require all 2^N subset scenarios).
    The gap between the cumulative stack-top and the dashed line is the
    nonlinear interaction residual — visible by construction.

    Args:
        runs:        dict {name: {'pe': pe_result, 'stats': stats}} from
                     run_scenarios.  Must contain `baseline_name` and at
                     least one other (counterfactual) entry.
        model:       HANCHousingModel — used for par.a_grid.
        year0:       calendar year of t=0 (default 1992).
        year_end:    last year to plot (default 2024).
        baseline_name: key of the baseline scenario (default 'baseline').
        normalize:   y-axis units. 'level' (default) plots krone changes with
                     the hand-tuned level limits; 'pct_ss' divides each panel by
                     that stock's 1992 SS level (panels comparable in growth %,
                     but the urban/rural level gap is hidden, and a tiny rural
                     base inflates %); 'pct_gdp' divides every panel by one
                     common 1992 GDP (unit-free, regions stay comparable in
                     economic weight). Scaling is by a single constant per panel
                     so the additive stacking and residual gap are preserved.
        colors:      band colors for the non-baseline scenarios. None (default)
                     uses the tab10 cycle; a list/tuple is applied in scenario
                     order (cycled if shorter); a dict {scenario_name: color}
                     sets colors by name with a tab10 fallback for missing keys.
        legend_labels: optional {name: display_text} to rename legend entries
                     (keys are scenario names; `baseline_name` may also be a key
                     to relabel the dashed baseline line).
        legend_loc:  matplotlib legend `loc` (default 'best').
        legend_bbox: optional bbox_to_anchor (axes coords) to reposition the
                     legend; e.g. (-0.05, 1.0) nudges it left of the panel. When
                     set with legend_loc='best', loc falls back to 'upper left'.
        data_file:   path to datagraphs.xlsx; its 'Urban/Rural housing wealth'
                     columns are overlaid (as 'data') on the two housing panels,
                     indexed to the model's 1992 stock so only the data's growth
                     enters — comparable in every `normalize` mode.
        show_data:   set False to suppress the empirical overlay.
        figsize:     figure size (default (13, 9)).
        title:       optional figure suptitle (default auto-generated).

    Returns the matplotlib Figure.
    """
    import matplotlib.pyplot as plt

    if baseline_name not in runs:
        raise KeyError(f"baseline_name='{baseline_name}' not in runs "
                       f"(keys: {list(runs.keys())})")
    fan_names = [n for n in runs if n != baseline_name]
    if not fan_names:
        raise ValueError("Need at least one non-baseline scenario for "
                          "the decomposition.")

    # --- y-axis units: levels, % of each stock's 1992 SS, or % of 1992 GDP ---
    # Every layer here is a difference of levels, so scaling each panel by a
    # SINGLE constant (the 1992 SS level, or one common 1992 GDP) only relabels
    # the axis: the additive stacking and the nonlinear-residual gap are
    # preserved.  (Do NOT normalise period-by-period — that would distort it.)
    normalize = (normalize or 'level').lower()
    if normalize not in ('level', 'pct_ss', 'pct_gdp'):
        raise ValueError("normalize must be 'level', 'pct_ss' or 'pct_gdp'; "
                         f"got {normalize!r}.")

    gdp0 = np.nan
    if normalize == 'pct_gdp':
        ss = model.ss
        _Y   = float(getattr(ss, 'Y',   np.nan))
        _qu  = float(getattr(ss, 'q_u', np.nan)); _IHu = float(getattr(ss, 'IH_u', np.nan))
        _qr  = float(getattr(ss, 'q_r', np.nan)); _IHr = float(getattr(ss, 'IH_r', np.nan))
        gdp0 = _Y + _qu * _IHu + _qr * _IHr     # eq. 3.26: Ytilde + q^u IH^u + q^r IH^r
        if not np.isfinite(gdp0) or gdp0 <= 0.0:
            raise ValueError(
                "Could not compute a positive 1992 GDP from model.ss "
                "(need Y, q_u, q_r, IH_u, IH_r). Use normalize='pct_ss' or 'level'.")

    def _norm_factor(base_lvl_panel):
        """Single multiplicative constant for this panel (see note above)."""
        if normalize == 'level':
            return 1.0
        if normalize == 'pct_gdp':
            return 100.0 / gdp0
        base0 = float(base_lvl_panel[0])        # 1992 SS level of this stock
        if not np.isfinite(base0) or abs(base0) < 1e-12:
            return np.nan                        # tiny base → blanks the panel
        return 100.0 / base0

    if normalize == 'level':
        ylab, change_word = '', f'Δ from {year0}'
    elif normalize == 'pct_ss':
        ylab, change_word = '% of 1992 stock', f'% from {year0}'
    else:
        ylab, change_word = '% of 1992 GDP', f'% from {year0}'

    def _stocks(pe):
        """Return dict of the four wealth-stock series for one pe_result."""
        for k in ('D', 'H_u_hh', 'H_r_hh', 'q_u_path', 'q_r_path'):
            if k not in pe:
                raise KeyError(
                    f"pe_result missing '{k}' — re-run simulate_olg_pe.")
        D = np.asarray(pe['D'], dtype=float)
        H_u_hh = np.asarray(pe['H_u_hh'], dtype=float).ravel()
        H_r_hh = np.asarray(pe['H_r_hh'], dtype=float).ravel()
        q_u = np.asarray(pe['q_u_path'], dtype=float).ravel()
        q_r = np.asarray(pe['q_r_path'], dtype=float).ravel()
        a_grid = np.asarray(model.par.a_grid, dtype=float).ravel()
        T = D.shape[0]
        if 'hh_scale' in pe:
            hh_scale = np.asarray(pe['hh_scale'], dtype=float).ravel()[:T]
        else:
            # Fallback for stale pe_results predating the hh_scale storage.
            # Only correct if model.path.hh_scale matches this scenario's
            # path — re-run simulate_olg_pe / run_scenarios to be sure.
            hh_scale = np.asarray(_get_hh_level_scale_path(model),
                                  dtype=float).ravel()[:T]
        A_urban = ((D[:, :, [1, 3], :].sum(axis=(1, 2)) * a_grid[None, :])
                   .sum(axis=1) * hh_scale)
        A_rural = ((D[:, :, [0, 2], :].sum(axis=(1, 2)) * a_grid[None, :])
                   .sum(axis=1) * hh_scale)
        HW_urban = q_u * H_u_hh
        HW_rural = q_r * H_r_hh
        return {'A_urban': A_urban, 'A_rural': A_rural,
                'HW_urban': HW_urban, 'HW_rural': HW_rural}

    stocks = {name: _stocks(runs[name]['pe']) for name in runs}

    T = len(stocks[baseline_name]['A_urban'])
    years = year0 + np.arange(T)
    mask = (years >= year0) & (years <= year_end)
    yrs = years[mask]

    panels = [
        ('A_urban',  r'Urban liquid wealth $A^u$'),
        ('A_rural',  r'Rural liquid wealth $A^r$'),
        ('HW_urban', r'Urban housing wealth $q^u H^u$'),
        ('HW_rural', r'Rural housing wealth $q^r H^r$'),
    ]

    fig, axes = plt.subplots(2, 2, figsize=figsize, sharex=True)
    # Resolve a band color per non-baseline scenario. `colors` may be:
    #   None -> matplotlib tab10 cycle (by scenario order)
    #   list/tuple -> applied to fan_names in order (cycled if shorter)
    #   dict -> {scenario_name: color}, with tab10 fallback for any missing key
    _cycle = plt.cm.tab10.colors
    if colors is None:
        band_colors = [_cycle[i % len(_cycle)] for i in range(len(fan_names))]
    elif isinstance(colors, dict):
        band_colors = [colors.get(n, _cycle[i % len(_cycle)])
                       for i, n in enumerate(fan_names)]
    else:
        _seq = list(colors)
        band_colors = [_seq[i % len(_seq)] for i in range(len(fan_names))]

    # Empirical regional housing-wealth stocks via the shared loader in plots.py,
    # mapped to the housing-panel keys.  Missing file/columns -> no overlay.
    emp_hw = {}
    if show_data:
        try:
            import plots as _plots
            _emp = _plots.load_datagraphs(data_file, start_year=year0)
            if 'Urban housing wealth' in _emp:
                emp_hw['HW_urban'] = _emp['Urban housing wealth']
            if 'Rural housing wealth' in _emp:
                emp_hw['HW_rural'] = _emp['Rural housing wealth']
        except Exception:
            emp_hw = {}
    data_plotted = False

    for ax, (key, panel_title) in zip(axes.ravel(), panels):
        base_lvl = stocks[baseline_name][key][mask]
        f = _norm_factor(base_lvl)                          # panel scaling const
        base0_raw = float(base_lvl[0])                      # model 1992 stock (unscaled)
        base_lvl = base_lvl * f
        base_change = base_lvl - base_lvl[0]                # Δ (or %) from year0

        # Stack contributions one scenario at a time.  contribution_X(t) =
        # base_change(t) - scen_X_change(t) = base_lvl(t) - scen_X_lvl(t)
        # (since baseline and counterfactuals share the year0 initial value).
        # Positives stack upward from 0, negatives stack downward from 0, so
        # each scenario's contribution is visible whichever side it lands on
        # — and the layers never overlap.
        pos_cum = np.zeros_like(yrs, dtype=float)
        neg_cum = np.zeros_like(yrs, dtype=float)
        for i, n in enumerate(fan_names):
            scen_lvl = stocks[n][key][mask] * f
            contribution = base_lvl - scen_lvl
            pos = np.where(contribution > 0.0, contribution, 0.0)
            neg = np.where(contribution < 0.0, contribution, 0.0)
            color = band_colors[i]
            # Thin white edges make the stack boundaries explicit so adjacent
            # same-hue bands (e.g. blue then green) don't look continuous.
            ax.fill_between(yrs, pos_cum, pos_cum + pos,
                            facecolor=color, alpha=0.85,
                            edgecolor='black', linewidth=0.5,
                            label=(legend_labels or {}).get(n, n))
            ax.fill_between(yrs, neg_cum, neg_cum + neg,
                            facecolor=color, alpha=0.85,
                            edgecolor='black', linewidth=0.5)
            pos_cum = pos_cum + pos
            neg_cum = neg_cum + neg

        # Baseline change overlay + zero reference.
        ax.plot(yrs, base_change, 'k-', lw=2,
                label=(legend_labels or {}).get(
                    baseline_name, f'{baseline_name} ({change_word})'))
        ax.axhline(0.0, color='gray', lw=0.5)

        # Empirical overlay (housing panels): index the data to the model's 1992
        # stock so only the data's GROWTH enters, then apply the same panel
        # factor f.  In pct_ss this reduces to 100*(data_t/data_1992 - 1).
        if key in emp_hw:
            d_years, d_vals = emp_hw[key]
            _d0 = d_vals[d_years == year0]
            d0 = float(_d0[0]) if (_d0.size and np.isfinite(_d0[0])) else np.nan
            if not np.isfinite(d0):
                _fin = d_vals[np.isfinite(d_vals)]
                d0 = float(_fin[0]) if _fin.size else np.nan
            if np.isfinite(d0) and d0 != 0.0:
                d_sel = (d_years >= year0) & (d_years <= year_end)
                d_change = f * base0_raw * (d_vals[d_sel] / d0 - 1.0)
                ax.plot(d_years[d_sel], d_change, color='k', lw=1.8, ls='--',
                        label='data')
                data_plotted = True

        ax.set_title(panel_title)
        ax.set_xlabel('year')
        if ylab:                                       # left and right columns
            ax.set_ylabel(ylab)
        ax.set_xlim(year0, year_end)
        ax.set_box_aspect(1.0)

    for _ax in (axes[0, 0], axes[0, 1]):               # year ticks on top row too
        _ax.tick_params(axis='x', labelbottom=True)

    _leg_kw = dict(fontsize=8, frameon=False, loc=legend_loc)
    if legend_bbox is not None:
        # bbox_to_anchor moves the legend; x<0 nudges it left of the axes.
        _leg_kw['bbox_to_anchor'] = legend_bbox
        if legend_loc == 'best':
            _leg_kw['loc'] = 'upper left'
    _handles, _labels = axes[0, 0].get_legend_handles_labels()
    if data_plotted:
        # The data line is only on the housing panels, so add a proxy entry.
        from matplotlib.lines import Line2D
        _handles.append(Line2D([0], [0], color='k', lw=1.5, ls='--'))
        _labels.append((legend_labels or {}).get('data', 'data'))
    axes[0, 0].legend(_handles, _labels, **_leg_kw)

    if normalize == 'level':
        # Hand-tuned level limits (krone units); only valid for the level view.
        fig.axes[0].set_ylim(-15.0, 25.0)
        fig.axes[1].set_ylim(-15.0, 25.0)
        fig.axes[2].set_ylim(-2.0, 25.0)      # urban housing only
        fig.axes[3].set_ylim(-2.0, 10.0)       # rural housing only
        fig.tight_layout()                # re-snap layout after ylim change
    # For pct_ss / pct_gdp the panels autoscale (the % ranges differ by panel).

    if title is None:
        _suffix = {'level': '', 'pct_ss': ' — % of 1992 stock',
                   'pct_gdp': ' — % of 1992 GDP'}[normalize]
        title = f'Region-wise wealth decomposition ({year0}-{year_end}){_suffix}'
    fig.suptitle(title)
    fig.tight_layout()
    return fig

